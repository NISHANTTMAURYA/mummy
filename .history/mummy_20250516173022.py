import customtkinter as ctk
import os
import webbrowser
import shutil
import openpyxl
import tkinter as tk
from tkinter import ttk
from copy import copy

class ExcelPage(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent)
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        
        # Set cute color scheme based on appearance mode
        self.update_colors()
        
        # Create a cute container
        container = ctk.CTkFrame(self, fg_color=self.colors["card_bg"], corner_radius=20)
        container.grid(row=0, column=0, padx=80, pady=80, sticky="nsew")
        container.grid_columnconfigure(0, weight=1)
        container.grid_rowconfigure(3, weight=1)
        
        # Add cute title
        title_frame = ctk.CTkFrame(container, fg_color="transparent")
        title_frame.grid(row=0, column=0, padx=30, pady=(30, 20), sticky="ew")
        
        ctk.CTkLabel(
            title_frame,
            text="✨ Excel Viewer ✨",
            font=ctk.CTkFont(size=32, weight="bold", family="Arial"),
            text_color="#ffffff"  # Always white for better visibility
        ).pack()
        
        # Add cute subtitle
        ctk.CTkLabel(
            container,
            text="Click the button below to open your Excel file",
            font=ctk.CTkFont(size=18, family="Arial"),
            text_color="#ffffff"  # Always white for better visibility
        ).grid(row=1, column=0, padx=30, pady=(0, 30))
        
        # Big cute Excel icon
        ctk.CTkLabel(
            container,
            text="📊",
            font=ctk.CTkFont(size=80)
        ).grid(row=2, column=0, pady=20)
        
        # Cute button with better styling
        self.open_button = ctk.CTkButton(
            container,
            text="📂 Open Excel File",
            command=self.open_excel_file,
            font=ctk.CTkFont(size=20, weight="bold", family="Arial"),
            fg_color=self.colors["accent"],
            hover_color=self.colors["accent_hover"],
            corner_radius=15,
            width=250,
            height=60,
            text_color="#ffffff"  # Always white for better visibility
        )
        self.open_button.grid(row=3, column=0, padx=40, pady=40)
        
        # Decorative bottom row
        bottom_frame = ctk.CTkFrame(container, fg_color="transparent")
        bottom_frame.grid(row=4, column=0, pady=(0, 20))
        
        for i, emoji in enumerate(["💕", "📈", "💕"]):
            ctk.CTkLabel(
                bottom_frame,
                text=emoji,
                font=ctk.CTkFont(size=24)
            ).grid(row=0, column=i, padx=15)

    def update_colors(self):
        """Set color scheme based on appearance mode"""
        if ctk.get_appearance_mode() == "Dark":
            self.colors = {
                "bg_primary": "#2d2438",  # Dark purple background
                "bg_secondary": "#332b40",  # Medium dark purple
                "card_bg": "#3a2b4a",  # Medium purple for cards
                "accent": "#b76edc",  # Bright purple accent
                "accent_hover": "#c78ae8",  # Lighter purple for hover
                "text_primary": "#e6e6e6",  # Light gray for text
                "dropdown_bg": "#3a2b4a",  # Medium purple for dropdown
                "dropdown_hover": "#473960",  # Slightly lighter purple for hover
                "input_bg": "#3a2b4a",  # Medium purple for input
                "border": "#b76edc",  # Bright purple for borders
                "file_bg": "#473960",  # Light purple for file rows
                "file_hover": "#524372",  # Lighter purple for file row hover
                "tree_bg": "#251f30",  # Darker purple for tree
                "tree_even": "#2d2438",  # Dark purple
                "tree_odd": "#332b40",  # Medium dark purple
                "tree_header": "#b76edc"  # Bright purple for headers
            }
        else:
            self.colors = {
                "bg_primary": "#fff5f9",  # Very light pink background
                "bg_secondary": "#fff0f5",  # Light pink
                "card_bg": "#ffebf2",  # Lighter pink for cards
                "accent": "#ffacc7",  # Medium pink accent
                "accent_hover": "#ff85a1",  # Darker pink for hover
                "text_primary": "#4a4a4a",  # Dark gray for text
                "dropdown_bg": "#ffebf2",  # Lighter pink for dropdown
                "dropdown_hover": "#ffd6e0",  # Medium light pink for hover
                "input_bg": "#ffebf2",  # Lighter pink for input
                "border": "#ffacc7",  # Medium pink for borders
                "file_bg": "#ffd6e0",  # Medium light pink for file rows
                "file_hover": "#ffc1d5",  # Slightly darker pink for file row hover
                "tree_bg": "#fff5f9",  # Very light pink for tree
                "tree_even": "#fff0f5",  # Light pink
                "tree_odd": "#ffebf2",  # Lighter pink
                "tree_header": "#ffacc7"  # Medium pink for headers
            }

    def open_excel_file(self):
        excel_path = os.path.abspath("iso_excel.xlsx")
        webbrowser.open(f"file://{excel_path}")

class EditPage(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent)
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(4, weight=1)
        
        # Set cute color scheme based on appearance mode
        self.update_colors()
        
        # Header frame with cute styling
        header_frame = ctk.CTkFrame(self, fg_color=self.colors["card_bg"], corner_radius=15)
        header_frame.grid(row=0, column=0, columnspan=2, sticky="ew", padx=40, pady=(30, 20))
        header_frame.grid_columnconfigure(1, weight=1)
        
        # Title with cute emoji
        title_label = ctk.CTkLabel(
            header_frame, 
            text="📝 Edit Excel Data", 
            font=ctk.CTkFont(size=22, weight="bold", family="Arial"),
            text_color="#ffffff"  # Always white for better visibility
        )
        title_label.grid(row=0, column=0, columnspan=2, sticky="w", padx=20, pady=(15, 20))
        
        # File selection with cute styling
        file_select_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        file_select_frame.grid(row=1, column=0, columnspan=2, sticky="ew", padx=20, pady=(0, 10))
        file_select_frame.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(
            file_select_frame, 
            text="🗂️ Select Excel File:", 
            font=ctk.CTkFont(size=16, weight="bold", family="Arial"),
            text_color="#ffffff"  # Always white for better visibility
        ).grid(row=0, column=0, sticky="w", pady=(0, 10))
        
        self.file_var = ctk.StringVar()
        self.file_menu = ctk.CTkOptionMenu(
            file_select_frame, 
            variable=self.file_var, 
            values=self.get_file_list(), 
            command=self.on_file_change, 
            width=400,
            height=40,
            font=ctk.CTkFont(size=14, family="Arial"),
            fg_color=self.colors["dropdown_bg"],
            button_color=self.colors["accent"],
            button_hover_color=self.colors["accent_hover"],
            dropdown_fg_color=self.colors["dropdown_bg"],
            dropdown_hover_color=self.colors["dropdown_hover"],
            corner_radius=10,
            text_color="#ffffff"  # Always white for better visibility
        )
        self.file_menu.grid(row=0, column=1, sticky="w", padx=(20, 0), pady=(0, 10))
        
        # Month selection with cute styling
        month_select_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        month_select_frame.grid(row=2, column=0, columnspan=2, sticky="ew", padx=20, pady=(10, 20))
        month_select_frame.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(
            month_select_frame, 
            text="📅 Select Month:", 
            font=ctk.CTkFont(size=16, weight="bold", family="Arial"),
            text_color="#ffffff"  # Always white for better visibility
        ).grid(row=0, column=0, sticky="w")
        
        self.month_var = ctk.StringVar()
        self.month_menu = ctk.CTkOptionMenu(
            month_select_frame, 
            variable=self.month_var, 
            values=[], 
            command=self.on_month_change, 
            width=200,
            height=40,
            font=ctk.CTkFont(size=14, family="Arial"),
            fg_color=self.colors["dropdown_bg"],
            button_color=self.colors["accent"],
            button_hover_color=self.colors["accent_hover"],
            dropdown_fg_color=self.colors["dropdown_bg"],
            dropdown_hover_color=self.colors["dropdown_hover"],
            corner_radius=10,
            text_color="#ffffff"  # Always white for better visibility
        )
        self.month_menu.grid(row=0, column=1, sticky="w", padx=(20, 0))
        
        # Refresh button with cute styling
        self.refresh_button = ctk.CTkButton(
            month_select_frame,
            text="🔄 Refresh",
            command=self.refresh_data,
            font=ctk.CTkFont(size=14, weight="bold", family="Arial"),
            width=120,
            height=40,
            fg_color=self.colors["accent"],
            hover_color=self.colors["accent_hover"],
            corner_radius=10,
            text_color="#ffffff"  # Always white for better visibility
        )
        self.refresh_button.grid(row=0, column=1, sticky="e", padx=(0, 10))
        
        # Data frame for the table with cute styling
        self.data_frame = ctk.CTkFrame(self, fg_color=self.colors["bg_secondary"], corner_radius=15)
        self.data_frame.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=40, pady=20)
        self.data_widgets = []
        
        # Save button with cute styling
        button_frame = ctk.CTkFrame(self, fg_color=self.colors["card_bg"], corner_radius=15)
        button_frame.grid(row=3, column=0, columnspan=2, sticky="ew", padx=40, pady=(10, 20))
        button_frame.grid_columnconfigure(0, weight=1)
        
        self.save_button = ctk.CTkButton(
            button_frame, 
            text="💾 Save Changes", 
            command=self.save_changes, 
            font=ctk.CTkFont(size=18, weight="bold", family="Arial"),
            height=50,
            width=220,
            fg_color=self.colors["accent"],
            hover_color=self.colors["accent_hover"],
            corner_radius=15,
            text_color="#ffffff"  # Always white for better visibility
        )
        self.save_button.grid(row=0, column=0, pady=15)
        
        # Status label with cute styling
        self.status_label = ctk.CTkLabel(
            button_frame, 
            text="", 
            font=ctk.CTkFont(size=14, family="Arial"),
            text_color="#ffffff"  # Will be set when displaying messages
        )
        self.status_label.grid(row=1, column=0, pady=(5, 15))
        
        self.current_file = None
        self.current_month = None
        self.month_col_ranges = {}
        self.sub_headers = []
        self.initials = []
        self._last_edited_cell = None
        self._refresh_pending = False
        self.load_files()

    def update_colors(self):
        """Set color scheme based on appearance mode"""
        if ctk.get_appearance_mode() == "Dark":
            self.colors = {
                "bg_primary": "#2d2438",  # Dark purple background
                "bg_secondary": "#332b40",  # Medium dark purple
                "card_bg": "#3a2b4a",  # Medium purple for cards
                "accent": "#b76edc",  # Bright purple accent
                "accent_hover": "#c78ae8",  # Lighter purple for hover
                "text_primary": "#e6e6e6",  # Light gray for text
                "dropdown_bg": "#3a2b4a",  # Medium purple for dropdown
                "dropdown_hover": "#473960",  # Slightly lighter purple for hover
                "input_bg": "#3a2b4a",  # Medium purple for input
                "border": "#b76edc",  # Bright purple for borders
                "file_bg": "#473960",  # Light purple for file rows
                "file_hover": "#524372",  # Lighter purple for file row hover
                "tree_bg": "#251f30",  # Darker purple for tree
                "tree_even": "#2d2438",  # Dark purple
                "tree_odd": "#332b40",  # Medium dark purple
                "tree_header": "#b76edc"  # Bright purple for headers
            }
        else:
            self.colors = {
                "bg_primary": "#fff5f9",  # Very light pink background
                "bg_secondary": "#fff0f5",  # Light pink
                "card_bg": "#ffebf2",  # Lighter pink for cards
                "accent": "#ffacc7",  # Medium pink accent
                "accent_hover": "#ff85a1",  # Darker pink for hover
                "text_primary": "#4a4a4a",  # Dark gray for text
                "dropdown_bg": "#ffebf2",  # Lighter pink for dropdown
                "dropdown_hover": "#ffd6e0",  # Medium light pink for hover
                "input_bg": "#ffebf2",  # Lighter pink for input
                "border": "#ffacc7",  # Medium pink for borders
                "file_bg": "#ffd6e0",  # Medium light pink for file rows
                "file_hover": "#ffc1d5",  # Slightly darker pink for file row hover
                "tree_bg": "#fff5f9",  # Very light pink for tree
                "tree_even": "#fff0f5",  # Light pink
                "tree_odd": "#ffebf2",  # Lighter pink
                "tree_header": "#ffacc7"  # Medium pink for headers
            }

    def get_file_list(self):
        """Get a filtered list of Excel files for the file selection dropdown."""
        if not os.path.exists("excel_copies"):
            return []
            
        # Get only valid Excel files, filtering out system files and temp files
        all_files = os.listdir("excel_copies")
        return [f for f in all_files if (
            f.endswith(".xlsx") and  # Only Excel files
            not f.startswith("~") and  # Not temp files
            not f.startswith("$") and  # Not system files
            not f.startswith(".")  # Not hidden files
        )]

    def load_files(self):
        files = self.get_file_list()
        self.file_menu.configure(values=files)
        if files:
            self.file_var.set(files[0])
            self.on_file_change(files[0])
        else:
            self.file_var.set("")
            self.month_menu.configure(values=[])
            self.clear_data_frame()

    def on_file_change(self, filename):
        import openpyxl
        self.current_file = os.path.join("excel_copies", filename)
        
        # Extract standard from filename if present
        self.current_std = "Unknown"
        if "_FYJC.xlsx" in filename:
            self.current_std = "FYJC (11th)"
        elif "_SYJC.xlsx" in filename:
            self.current_std = "SYJC (12th)"
        
        try:
            # When parsing the file structure, use normal mode to get formulas
            wb = openpyxl.load_workbook(self.current_file)
            ws = wb.active
            
            # First, find the months and their column ranges by analyzing the first row
            months = []
            month_col_ranges = {}
            
            # Get all cells from the first row
            first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            second_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
            
            # Scan columns to find month headers
            current_month = None
            start_col = None
            
            for idx, val in enumerate(first_row):
                if idx < 2:  # Skip SR.NO. and INITIALS columns
                    continue
                    
                # If we find a month or TOTAL header
                if val and isinstance(val, str) and val.strip() and val.strip().upper() not in ("SR.NO.", "INITIALS"):
                    # If we were tracking a previous month, save its range
                    if current_month:
                        month_col_ranges[current_month] = (start_col, idx-1)
                    
                    # Start tracking new month
                    current_month = str(val).strip()
                    # Skip adding TOTAL to the dropdown
                    if current_month.upper() != "TOTAL":
                        months.append(current_month)
                    start_col = idx
            
            # Save the last month's range
            if current_month:
                month_col_ranges[current_month] = (start_col, len(first_row)-1)
            
            # Store the column ranges and headers
            self.month_col_ranges = month_col_ranges
            self.sub_headers = second_row  # Second row contains column headers like ALOTTED, E-Act, etc.
            
            # Get all initials (skip TOTAL/empty)
            initials = []
            for row in ws.iter_rows(min_row=3, max_col=2, values_only=True):
                if row[1] and str(row[1]).strip() and str(row[1]).strip().upper() != "TOTAL":
                    initials.append(str(row[1]).strip())
            
            self.initials = initials
            
            # Filter out TOTAL from the dropdown
            filtered_months = [m for m in months if m.upper() != "TOTAL"]
            self.month_menu.configure(values=filtered_months)
            
            if filtered_months:
                self.month_var.set(filtered_months[0])
                self.on_month_change(filtered_months[0])
            else:
                self.month_var.set("")
                self.clear_data_frame()
                
        except Exception as e:
            self.status_label.configure(text=f"Error loading file: {e}", text_color="red")
            self.month_menu.configure(values=[])
            self.clear_data_frame()

    def on_month_change(self, month):
        """Handle month selection change"""
        self.current_month = month
        # Clear any existing edit widgets
        if hasattr(self, 'edit_entry') and self.edit_entry:
            edit_frame, _, _, _, _ = self.edit_entry
            if edit_frame.winfo_exists():
                edit_frame.destroy()
            self.edit_entry = None
        # Display data for the selected month
        self.display_data()

    def clear_data_frame(self):
        for widget in self.data_frame.winfo_children():
            widget.destroy()
        self.data_widgets = []
        self.tree = None
        self.tree_vars = {}

    def display_data(self):
        self.clear_data_frame()
        if not (self.current_file and self.current_month):
            return
            
        import openpyxl
        try:
            # Update colors in case appearance mode changed
            self.update_colors()
            
            # Unbind previous resize handlers
            try:
                self.winfo_toplevel().unbind("<Configure>")
            except:
                pass
                
            # Load workbook with data_only=True to get calculated values instead of formulas
            wb = openpyxl.load_workbook(self.current_file, data_only=True)
            ws = wb.active
            
            # Get the column range for the selected month
            col_range = self.month_col_ranges.get(self.current_month)
            if not col_range:
                self.status_label.configure(text="Month not found.", text_color="red")
                ctk.CTkLabel(self.data_frame, text="No data found for selected month.", font=ctk.CTkFont(size=14)).grid(row=0, column=0, padx=10, pady=10)
                return
                
            start_col, end_col = col_range
            
            # Define which headers we want to display/edit
            editable_headers = ["ALOTTED", "E-Act", "E-Add"]
            headers = ["Initial"]
            header_indices = []
            
            # Find the indices of editable columns within the month's range
            for col in range(start_col, end_col+1):
                if col < len(self.sub_headers):  # Ensure we don't go out of bounds
                    header = self.sub_headers[col]
                    if header in editable_headers:
                        headers.append(header)
                        header_indices.append(col)
            
            # Gather all initials and their values for the selected month
            data = []
            row_indices = []
            
            for i, row in enumerate(ws.iter_rows(min_row=3, max_col=2, values_only=True), start=3):
                initial = row[1]
                if initial and str(initial).strip() and str(initial).strip().upper() != "TOTAL":
                    values = [initial]
                    
                    for col, h in zip(header_indices, headers[1:]):
                        # Ensure we don't try to access columns outside worksheet bounds
                        if col+1 <= ws.max_column:
                            cell = ws.cell(row=i, column=col+1)
                            val = cell.value
                            
                            if h == "E-Add":
                                if val == '+' or val == '+ ' or val is None:
                                    val = ''
                                elif val is not None:
                                    # Clean up the value - keep only the number part
                                    val = str(val).strip().replace('+', '').strip()
                            elif val is None or str(val).strip().lower() == 'none':
                                val = ''
                            else:
                                # Convert to string to display numbers properly
                                val = str(val)
                        else:
                            val = ''
                            
                        values.append(val)
                    
                    data.append(values)
                    row_indices.append(i)
                    
            if not headers or not data:
                ctk.CTkLabel(self.data_frame, text="No data to display for this selection.", font=ctk.CTkFont(size=14)).grid(row=0, column=0, padx=10, pady=10)
                return

            # Add header info above the table
            header_frame = ctk.CTkFrame(self.data_frame, fg_color=self.colors["card_bg"], corner_radius=10)
            header_frame.grid(row=0, column=0, sticky="new", padx=20, pady=(10, 20), ipady=10)
            header_frame.grid_columnconfigure(0, weight=1)
            
            # Add file info (year, term, standard)
            info_text = f"🗓️ {self.current_month} Data"
            
            # Extract year from filename if possible
            year_text = ""
            if self.current_file:
                filename = os.path.basename(self.current_file)
                if "_202" in filename:  # Look for year pattern
                    year_part = filename.split("_")[2]  # Assuming format iso_excel_YYYY-YYYY_...
                    if year_part and "-" in year_part:
                        year_text = f" • {year_part}"
            
            # Add standard info if available
            std_text = ""
            if hasattr(self, 'current_std') and self.current_std != "Unknown":
                std_name = "11th Standard" if "FYJC" in self.current_std else "12th Standard"
                std_text = f" • {std_name}"
            
            # Combine all info in a single clean label
            full_info = info_text + year_text + std_text
            
            ctk.CTkLabel(
                header_frame,
                text=full_info,
                font=ctk.CTkFont(size=22, weight="bold"),
                text_color=self.colors["accent"]
            ).grid(row=0, column=0, padx=15, pady=5)

            # Create a custom frame for the table with border
            table_container = ctk.CTkFrame(self.data_frame, fg_color="transparent")
            table_container.grid(row=1, column=0, sticky="nsew", padx=20, pady=0)
            table_container.grid_columnconfigure(0, weight=1)
            table_container.grid_rowconfigure(0, weight=1)
            
            # Add a border frame around the table with cute styling
            border_frame = ctk.CTkFrame(table_container, fg_color=self.colors["tree_bg"], border_width=3, border_color=self.colors["border"], corner_radius=15)
            border_frame.grid(row=0, column=0, sticky="nsew", padx=0, pady=0)
            border_frame.grid_columnconfigure(0, weight=1)
            border_frame.grid_rowconfigure(0, weight=1)
            
            # Configure style for the treeview with cute styling
            style = ttk.Style()
            style.theme_use("clam")  # Use clam theme which supports more customization
            
            # Configure the Treeview colors and font
            style.configure("Cute.Treeview", 
                background=self.colors["tree_bg"],  # Match theme background
                foreground=self.colors["text_primary"],  # Match theme text color
                rowheight=50,  # Increased row height
                fieldbackground=self.colors["tree_bg"],
                borderwidth=0)  # Hide border
            
            # Configure the header style with cute styling
            style.configure("Cute.Treeview.Heading",
                background=self.colors["tree_header"],  # Match theme accent
                foreground="white",
                relief="raised",
                borderwidth=0,
                font=('Arial', 16, 'bold'))  # Increased font size and bold
            
            # Configure selection colors
            style.map('Cute.Treeview', 
                background=[('selected', self.colors["accent"])],  # Match theme accent
                foreground=[('selected', 'white')])
            
            # Configure the Treeview to show grid lines
            style.layout("Cute.Treeview", [
                ('Cute.Treeview.treearea', {'sticky': 'nswe'})
            ])
            style.configure("Cute.Treeview", 
                            borderwidth=0,
                            relief="flat")
                
            # Create Treeview with increased row height and cute styling
            self.tree = ttk.Treeview(border_frame, columns=headers, show="headings", 
                                    height=len(data), style="Cute.Treeview")
            
            # Configure column widths and headings
            column_width = 180
            for h in headers:
                self.tree.heading(h, text=h)
                self.tree.column(h, width=column_width, anchor="center")
            
            # Add data to the treeview with alternating row colors
            for idx, row in enumerate(data):
                item_id = self.tree.insert("", "end", values=row)
                
                # Apply alternating row colors
                if idx % 2 == 1:
                    self.tree.item(item_id, tags=('odd_row',))
                else:
                    self.tree.item(item_id, tags=('even_row',))
            
            # Configure row styles with cute alternating colors
            self.tree.tag_configure('odd_row', background=self.colors["tree_odd"], font=('Arial', 14))
            self.tree.tag_configure('even_row', background=self.colors["tree_even"], font=('Arial', 14))
            
            # Add horizontal grid lines tag
            self.tree.tag_configure('bottom_line', background=self.colors["border"])
            
            # Add a custom scrollbar with cute styling
            scrollbar_style = ttk.Style()
            scrollbar_style.configure("Cute.Vertical.TScrollbar", 
                                      background=self.colors["accent"], 
                                      troughcolor=self.colors["tree_bg"],
                                      bordercolor=self.colors["border"],
                                      arrowcolor="white")
            
            scrollbar = ttk.Scrollbar(border_frame, orient="vertical", 
                                      command=self.tree.yview,
                                      style="Cute.Vertical.TScrollbar")
            self.tree.configure(yscrollcommand=scrollbar.set)
            
            # Place the treeview and scrollbar
            self.tree.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
            scrollbar.grid(row=0, column=1, sticky="ns", pady=5)
            
            # Configure grid for proper layout
            self.data_frame.grid_columnconfigure(0, weight=1)
            self.data_frame.grid_rowconfigure(1, weight=1)  # The row with the table should expand
            
            # Create a custom entry widget for editing cells
            self.edit_entry = None
            
            # Bind the double-click event using a dedicated method
            self.tree.bind("<Double-1>", self.on_double_click)
            
            # Store for saving
            self.data_widgets = [(headers, row_indices, header_indices)]
            self.status_label.configure(text="", text_color="green")
            
        except Exception as e:
            self.status_label.configure(text=f"Error: {e}", text_color="red")
            ctk.CTkLabel(self.data_frame, text=f"Error: {e}", font=ctk.CTkFont(size=14)).grid(row=0, column=0, padx=10, pady=10)

    def _save_single_cell(self, item_id, col_idx, header, value):
        """Save a single cell value directly to Excel without reloading the entire sheet"""
        try:
            # Get the row index from the data_widgets
            if not self.data_widgets:
                return False
                
            headers, row_indices, header_indices = self.data_widgets[0]
            
            # Get the item index in the treeview
            item_index = self.tree.index(item_id)
            if item_index >= len(row_indices):
                return False
                
            # Get the Excel row and column
            row_idx = row_indices[item_index]
            col = header_indices[col_idx - 1]  # -1 because col_idx includes Initial column
            
            # Open the workbook (load with formulas for saving)
            import openpyxl
            wb = openpyxl.load_workbook(self.current_file)
            ws = wb.active
            
            # Process the value based on the header
            if header == "E-Add":
                # Process E-Add value
                v = str(value).strip() if value is not None else ""
                v = v.replace('"', '').replace("'", '').replace('+', '').strip()
                
                # Always save with a '+' prefix for E-Add values
                if v:
                    cell_value = f'+{v}'
                else:
                    cell_value = '+'
                
                # Get the actual cell
                cell = ws.cell(row=row_idx, column=col+1)
                # Set its value
                cell.value = cell_value
            else:
                # For other headers, just use the value as is
                cell_value = value if value and value.strip() else None
                ws.cell(row=row_idx, column=col+1).value = cell_value
            
            # Save the workbook
            wb.save(self.current_file)
            
            # Don't refresh when moving to next cell during rapid edits
            # We'll set a flag so we know this cell was just edited
            self._last_edited_cell = (item_id, col_idx)
            
            # Update status
            self.status_label.configure(text="Changes saved!", text_color="green")
            return True
        except Exception as e:
            self.status_label.configure(text=f"Error saving: {e}", text_color="red")
            return False

    def save_changes(self):
        """Save all changes to the Excel file."""
        if not self.current_file or not self.current_file.endswith(".xlsx"):
            self.status_label.configure(text="No file selected or not an Excel file.", text_color="red")
            return
        
        # First, check if we are currently editing a cell and save it if needed
        if hasattr(self, 'edit_entry') and self.edit_entry:
            edit_frame, entry, item, col_idx, header = self.edit_entry
            if edit_frame.winfo_exists():
                # Get the current value from the entry
                new_value = entry.get().strip()
                
                # Update the treeview
                values = list(self.tree.item(item, 'values'))
                old_value = values[col_idx]
                values[col_idx] = new_value
                self.tree.item(item, values=values)
                
                # Clean up
                edit_frame.destroy()
                self.edit_entry = None
                
                # If the value has changed, save it
                if old_value != new_value:
                    self._save_single_cell(item, col_idx, header, new_value)
        
        import openpyxl
        try:
            # Get the current values directly from the treeview
            all_items = self.tree.get_children()
            current_values = []
            for item in all_items:
                current_values.append(self.tree.item(item, 'values'))
            
            # Open the workbook
            wb = openpyxl.load_workbook(self.current_file)
            ws = wb.active
            
            # Get row indices from data_widgets
            if not self.data_widgets:
                self.status_label.configure(text="No data to save.", text_color="red")
                return
                
            headers, row_indices, header_indices = self.data_widgets[0]
            
            # Save each value
            for row_idx, values in zip(row_indices, current_values):
                for i, col in enumerate(header_indices):
                    header = headers[i+1]  # +1 because headers[0] is "Initial"
                    val = values[i+1]      # +1 because values[0] is Initial
                    
                    if header == "E-Add":
                        # Process E-Add value
                        v = str(val).strip() if val is not None else ""
                        v = v.replace('"', '').replace("'", '').replace('+', '').strip()
                        
                        # Always save with a '+' prefix for E-Add values
                        if v:
                            cell_value = f'+{v}'
                        else:
                            cell_value = '+'
                        
                        # Get the actual cell
                        cell = ws.cell(row=row_idx, column=col+1)
                        # Set its value
                        cell.value = cell_value
                    else:
                        # Process other values
                        if val is not None and str(val).strip() != '' and str(val).strip().lower() != 'none':
                            ws.cell(row=row_idx, column=col+1).value = val
                        else:
                            ws.cell(row=row_idx, column=col+1).value = None
            
            # Save the workbook
            wb.save(self.current_file)
            
            # Force reload of data after save all changes
            self.refresh_data()
            
            self.status_label.configure(text="All changes saved!", text_color="green")
            
        except Exception as e:
            self.status_label.configure(text=f"Error saving: {e}", text_color="red")

    def on_double_click(self, event):
        """Handle double-click on table cell"""
        if not hasattr(self, 'tree') or not self.tree:
            return
            
        item = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        if not item or not column:
            return
        
        # Get column index (skip if it's the Initial column)
        col_idx = int(column.replace('#','')) - 1
        if col_idx == 0:
            return
        
        # Get current value and position for edit
        x, y, width, height = self.tree.bbox(item, column)
        current_value = self.tree.item(item, 'values')[col_idx]
        
        # Get headers from data_widgets
        if not self.data_widgets:
            return
        headers = self.data_widgets[0][0]
        
        # For E-Add column, remove the '+' if present
        header = headers[col_idx]
        if header == "E-Add" and current_value and str(current_value).startswith('+'):
            current_value = str(current_value).replace('+', '').strip()
        
        # Create a frame for better control with cute styling
        edit_frame = tk.Frame(self.tree, bg=self.colors["accent"], highlightthickness=2, highlightbackground=self.colors["accent_hover"])
        edit_frame.place(x=x, y=y, width=width, height=height)
        
        # Create the entry widget with larger font and cute styling
        entry_var = tk.StringVar(value=current_value if current_value else "")
        entry = tk.Entry(edit_frame, textvariable=entry_var, 
                         font=('Arial', 16, 'bold'),  # Increased font size and bold
                         bg=self.colors["accent"],
                         fg="white",
                         bd=0,
                         highlightthickness=0,
                         justify=tk.CENTER)  # Center-aligned text
        entry.pack(fill="both", expand=True)
        entry.focus_set()
        entry.select_range(0, tk.END)
        
        # Store reference to edit widgets
        self.edit_entry = (edit_frame, entry, item, col_idx, header)
        
        def save_edit(event=None, move_to_next=True):
            # Get the new value
            new_value = entry_var.get().strip()
            
            # Update the treeview
            values = list(self.tree.item(item, 'values'))
            old_value = values[col_idx]
            values[col_idx] = new_value
            self.tree.item(item, values=values)
            
            # Clean up
            edit_frame.destroy()
            self.edit_entry = None
            
            # If the value hasn't changed, don't save
            save_success = True
            if old_value != new_value:
                # Save changes to Excel directly
                save_success = self._save_single_cell(item, col_idx, header, new_value)
            
            # Move to the next cell if requested
            if move_to_next and event and event.keysym == 'Return' and save_success:
                self._move_to_next_cell(item, col_idx)
        
        def cancel_edit(event=None):
            edit_frame.destroy()
            self.edit_entry = None
        
        # Bind events
        entry.bind("<Return>", save_edit)
        entry.bind("<Escape>", cancel_edit)
        entry.bind("<FocusOut>", lambda e: save_edit(e, move_to_next=False))
    
    def _move_to_next_cell(self, current_item, current_col_idx):
        """Move to the next editable cell after editing"""
        if not self.tree or not hasattr(self, 'tree'):
            return
            
        if not self.data_widgets:
            return
            
        headers = self.data_widgets[0][0]
        num_columns = len(headers)
        
        # If we're not at the last column, move to the next column in the same row
        if current_col_idx < num_columns - 1:
            next_col_idx = current_col_idx + 1
            # Skip the Initial column (0)
            if next_col_idx == 0:
                next_col_idx = 1
                
            # Create column identifier
            next_column = f"#{next_col_idx+1}"  # +1 because treeview columns are 1-indexed
            
            # Simulate a double-click on the next cell
            bbox = self.tree.bbox(current_item, next_column)
            if bbox:
                x, y, _, _ = bbox
                event = type('Event', (), {'x': x+5, 'y': y+5})  # +5 to ensure we're inside the cell
                self.on_double_click(event)
        else:
            # We're at the last column, move to the first column of the next row
            items = self.tree.get_children()
            try:
                current_index = items.index(current_item)
                if current_index < len(items) - 1:
                    next_item = items[current_index + 1]
                    # Move to the first editable column (skip Initial)
                    next_column = "#2"  # Column index 1 (2nd column) is the first editable one
                    
                    # Simulate a double-click on the next cell
                    bbox = self.tree.bbox(next_item, next_column)
                    if bbox:
                        x, y, _, _ = bbox
                        event = type('Event', (), {'x': x+5, 'y': y+5})  # +5 to ensure we're inside the cell
                        self.on_double_click(event)
            except ValueError:
                pass  # Item not found in the list

    def refresh_data(self):
        """Refresh the current data view"""
        if self.current_month:
            self.display_data()
            self.status_label.configure(text="Data refreshed!", text_color="green")

class CopyPage(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent)
        self.grid_columnconfigure(0, weight=1)
        # Adjust row configuration to give more space to the file list
        self.grid_rowconfigure(0, weight=0)  # Top section doesn't need to expand
        self.grid_rowconfigure(1, weight=1)  # Bottom section should expand to fill space
        
        # Set cute color scheme based on appearance mode
        self.update_colors()

        # --- Top: Create Copy Section ---
        top_frame = ctk.CTkFrame(self, fg_color=self.colors["card_bg"], corner_radius=20)
        top_frame.grid(row=0, column=0, sticky="ew", padx=40, pady=(30, 15))
        top_frame.grid_columnconfigure(0, weight=1)

        # Cute title with emoji
        title_frame = ctk.CTkFrame(top_frame, fg_color="transparent")
        title_frame.grid(row=0, column=0, sticky="ew", padx=20, pady=(20, 10))
        title_frame.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(title_frame, text="✨", font=ctk.CTkFont(size=28)).grid(row=0, column=0, padx=(0, 10))
        title = ctk.CTkLabel(title_frame, text="Create Yearly Excel Copy", 
                           font=ctk.CTkFont(size=24, weight="bold", family="Arial"),
                           text_color="#ffffff")  # Always white for better visibility
        title.grid(row=0, column=1, sticky="w")
        ctk.CTkLabel(title_frame, text="✨", font=ctk.CTkFont(size=28)).grid(row=0, column=2, padx=(10, 0))

        # Entry frame with cute styling
        entry_frame = ctk.CTkFrame(top_frame, fg_color=self.colors["bg_secondary"], corner_radius=15)
        entry_frame.grid(row=1, column=0, sticky="ew", pady=(10, 20), padx=20)
        entry_frame.grid_columnconfigure(1, weight=1)

        # Year entry with cute label
        year_label = ctk.CTkLabel(entry_frame, 
                                text="📅 Year Range:", 
                                font=ctk.CTkFont(size=16, weight="bold", family="Arial"),
                                text_color="#ffffff")  # Always white for better visibility
        year_label.grid(row=0, column=0, padx=20, pady=(20, 0), sticky="w")
        
        self.year_entry = ctk.CTkEntry(entry_frame, 
                                     placeholder_text="Enter year range (e.g. 2024-2025)", 
                                     font=ctk.CTkFont(size=16, family="Arial"), 
                                     width=300,
                                     height=40,
                                     fg_color=self.colors["input_bg"],
                                     border_color=self.colors["border"],
                                     corner_radius=10)
        self.year_entry.grid(row=1, column=0, padx=20, pady=(5, 20), sticky="w")
        
        # Add term selection dropdown with cute styling
        term_frame = ctk.CTkFrame(entry_frame, fg_color="transparent")
        term_frame.grid(row=2, column=0, padx=20, pady=(0, 20), sticky="w")
        
        ctk.CTkLabel(term_frame, 
                   text="📘 Select Term:", 
                   font=ctk.CTkFont(size=16, weight="bold", family="Arial"),
                   text_color="#ffffff").grid(row=0, column=0, sticky="w", padx=(0, 10))
        
        self.term_var = ctk.StringVar(value="term1")
        self.term_dropdown = ctk.CTkOptionMenu(term_frame,
                                             values=["term1", "term2"],
                                             variable=self.term_var,
                                             width=150,
                                             height=40,
                                             font=ctk.CTkFont(size=14, family="Arial"),
                                             fg_color=self.colors["dropdown_bg"],
                                             button_color=self.colors["accent"],
                                             button_hover_color=self.colors["accent_hover"],
                                             dropdown_fg_color=self.colors["dropdown_bg"],
                                             dropdown_hover_color=self.colors["dropdown_hover"],
                                             dropdown_font=ctk.CTkFont(size=14, family="Arial"),
                                             corner_radius=10,
                                             text_color="#ffffff")  # Always white for better visibility
        self.term_dropdown.grid(row=0, column=1, sticky="w", padx=(10, 0))
        
        # Add standard selection dropdown with cute styling
        std_frame = ctk.CTkFrame(entry_frame, fg_color="transparent")
        std_frame.grid(row=3, column=0, padx=20, pady=(0, 20), sticky="w")
        
        ctk.CTkLabel(std_frame, 
                   text="🎓 Select Standard:", 
                   font=ctk.CTkFont(size=16, weight="bold", family="Arial"),
                   text_color="#ffffff").grid(row=0, column=0, sticky="w", padx=(0, 10))
        
        self.std_var = ctk.StringVar(value="FYJC")
        self.std_dropdown = ctk.CTkOptionMenu(std_frame,
                                             values=["FYJC", "SYJC"],
                                             variable=self.std_var,
                                             width=150,
                                             height=40,
                                             font=ctk.CTkFont(size=14, family="Arial"),
                                             fg_color=self.colors["dropdown_bg"],
                                             button_color=self.colors["accent"],
                                             button_hover_color=self.colors["accent_hover"],
                                             dropdown_fg_color=self.colors["dropdown_bg"],
                                             dropdown_hover_color=self.colors["dropdown_hover"],
                                             dropdown_font=ctk.CTkFont(size=14, family="Arial"),
                                             corner_radius=10,
                                             text_color="#ffffff")  # Always white for better visibility
        self.std_dropdown.grid(row=0, column=1, sticky="w", padx=(10, 0))
        
        # Copy button with cute styling
        self.copy_button = ctk.CTkButton(entry_frame, 
                                       text="📋 Create Copy", 
                                       command=self.create_copy, 
                                       font=ctk.CTkFont(size=16, weight="bold", family="Arial"), 
                                       width=150,
                                       height=45,
                                       fg_color=self.colors["accent"],
                                       hover_color=self.colors["accent_hover"],
                                       corner_radius=15,
                                       text_color="#ffffff")  # Always white for better visibility
        self.copy_button.grid(row=1, column=1, rowspan=2, padx=20, pady=20, sticky="e")
        
        # Status label with cute styling
        self.status_label = ctk.CTkLabel(entry_frame, 
                                       text="", 
                                       font=ctk.CTkFont(size=14, family="Arial"),
                                       text_color="#ffffff")  # Will be changed when displaying messages
        self.status_label.grid(row=3, column=1, sticky="w", padx=20, pady=(0, 20))

        # --- Bottom: Available Copies List ---
        list_frame = ctk.CTkFrame(self, fg_color=self.colors["card_bg"], corner_radius=20)
        list_frame.grid(row=1, column=0, sticky="nsew", padx=40, pady=(20, 40))
        list_frame.grid_columnconfigure(0, weight=1)
        list_frame.grid_rowconfigure(2, weight=1)  # Row with files should expand

        # Title with decorative elements
        list_title_frame = ctk.CTkFrame(list_frame, fg_color="transparent")
        list_title_frame.grid(row=0, column=0, sticky="ew", padx=20, pady=(20, 5))
        list_title_frame.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(list_title_frame, text="📚", font=ctk.CTkFont(size=28)).grid(row=0, column=0, padx=(0, 10))
        list_title = ctk.CTkLabel(list_title_frame, 
                                text="Available Copies:", 
                                font=ctk.CTkFont(size=20, weight="bold", family="Arial"),
                                text_color="#ffffff")  # Always white for better visibility
        list_title.grid(row=0, column=1, sticky="w")
        
        # Add filter options
        filter_frame = ctk.CTkFrame(list_frame, fg_color=self.colors["bg_secondary"], corner_radius=10)
        filter_frame.grid(row=1, column=0, sticky="ew", padx=20, pady=(5, 10))
        filter_frame.grid_columnconfigure(4, weight=1)  # Push filters to the right
        
        # Create filter dropdowns
        filter_label = ctk.CTkLabel(filter_frame, 
                                  text="🔍 Filter by:", 
                                  font=ctk.CTkFont(size=14, weight="bold"),
                                  text_color="#ffffff")
        filter_label.grid(row=0, column=0, padx=(15, 5), pady=10, sticky="e")
        
        # Year filter
        self.year_filter_var = ctk.StringVar(value="All Years")
        self.year_filter = ctk.CTkOptionMenu(filter_frame,
                                           values=["All Years"],
                                           variable=self.year_filter_var,
                                           width=100,  # Reduced width
                                           height=30,
                                           font=ctk.CTkFont(size=12),
                                           fg_color=self.colors["dropdown_bg"],
                                           button_color=self.colors["accent"],
                                           button_hover_color=self.colors["accent_hover"],
                                           dropdown_fg_color=self.colors["dropdown_bg"],
                                           dropdown_hover_color=self.colors["dropdown_hover"],
                                           corner_radius=8,
                                           command=self.apply_filters)
        self.year_filter.grid(row=0, column=1, padx=(5, 5), pady=10, sticky="e")
        
        # Term filter
        self.term_filter_var = ctk.StringVar(value="All Terms")
        self.term_filter = ctk.CTkOptionMenu(filter_frame,
                                           values=["All Terms", "term1", "term2"],
                                           variable=self.term_filter_var,
                                           width=100,  # Reduced width
                                           height=30,
                                           font=ctk.CTkFont(size=12),
                                           fg_color=self.colors["dropdown_bg"],
                                           button_color=self.colors["accent"],
                                           button_hover_color=self.colors["accent_hover"],
                                           dropdown_fg_color=self.colors["dropdown_bg"],
                                           dropdown_hover_color=self.colors["dropdown_hover"],
                                           corner_radius=8,
                                           command=self.apply_filters)
        self.term_filter.grid(row=0, column=2, padx=(5, 5), pady=10, sticky="e")
        
        # Standard filter
        self.std_filter_var = ctk.StringVar(value="All Standards")
        self.std_filter = ctk.CTkOptionMenu(filter_frame,
                                           values=["All Standards", "FYJC", "SYJC"],
                                           variable=self.std_filter_var,
                                           width=100,  # Reduced width
                                           height=30,
                                           font=ctk.CTkFont(size=12),
                                           fg_color=self.colors["dropdown_bg"],
                                           button_color=self.colors["accent"],
                                           button_hover_color=self.colors["accent_hover"],
                                           dropdown_fg_color=self.colors["dropdown_bg"],
                                           dropdown_hover_color=self.colors["dropdown_hover"],
                                           corner_radius=8,
                                           command=self.apply_filters)
        self.std_filter.grid(row=0, column=3, padx=(5, 5), pady=10, sticky="e")
        
        # Reset button
        reset_btn = ctk.CTkButton(filter_frame,
                                 text="↺ Reset",
                                 width=70,  # Reduced width
                                 height=30,
                                 font=ctk.CTkFont(size=12, weight="bold"),
                                 fg_color=self.colors["accent"],
                                 hover_color=self.colors["accent_hover"],
                                 corner_radius=8,
                                 command=self.reset_filters)
        reset_btn.grid(row=0, column=5, padx=(5, 15), pady=10, sticky="e")

        # Create a container for the scrollable frame with proper border
        scroll_container = ctk.CTkFrame(list_frame, fg_color="transparent", border_width=2, 
                                      border_color=self.colors["border"], corner_radius=15)
        scroll_container.grid(row=2, column=0, sticky="nsew", padx=20, pady=(0, 10))
        scroll_container.grid_columnconfigure(0, weight=1)
        scroll_container.grid_rowconfigure(0, weight=1)
        
        # Scrollable frame for file list with cute styling
        self.scrollable_frame = ctk.CTkScrollableFrame(scroll_container, 
                                                     fg_color=self.colors["bg_secondary"], 
                                                     corner_radius=15,
                                                     border_width=0)
        self.scrollable_frame.grid(row=0, column=0, sticky="nsew", padx=0, pady=0)
        self.scrollable_frame.grid_columnconfigure(0, weight=1)
        
        # Decorative elements at the bottom
        deco_frame = ctk.CTkFrame(list_frame, fg_color="transparent")
        deco_frame.grid(row=3, column=0, sticky="ew", pady=(0, 10))
        
        for i, emoji in enumerate(["💗", "📊", "💗"]):
            ctk.CTkLabel(deco_frame, 
                       text=emoji, 
                       font=ctk.CTkFont(size=20)).grid(row=0, column=i, padx=20)
        
        # Initialize
        self.all_files = []
        self.refresh_file_list()
        self.new_file_path = None

    def update_colors(self):
        """Set color scheme based on appearance mode"""
        if ctk.get_appearance_mode() == "Dark":
            self.colors = {
                "bg_primary": "#2d2438",  # Dark purple background
                "bg_secondary": "#332b40",  # Medium dark purple
                "card_bg": "#3a2b4a",  # Medium purple for cards
                "accent": "#b76edc",  # Bright purple accent
                "accent_hover": "#c78ae8",  # Lighter purple for hover
                "text_primary": "#e6e6e6",  # Light gray for text
                "dropdown_bg": "#3a2b4a",  # Medium purple for dropdown
                "dropdown_hover": "#473960",  # Slightly lighter purple for hover
                "input_bg": "#3a2b4a",  # Medium purple for input
                "border": "#b76edc",  # Bright purple for borders
                "file_bg": "#473960",  # Light purple for file rows
                "file_hover": "#524372",  # Lighter purple for file row hover
                "tree_bg": "#251f30",  # Darker purple for tree
                "tree_even": "#2d2438",  # Dark purple
                "tree_odd": "#332b40",  # Medium dark purple
                "tree_header": "#b76edc"  # Bright purple for headers
            }
        else:
            self.colors = {
                "bg_primary": "#fff5f9",  # Very light pink background
                "bg_secondary": "#fff0f5",  # Light pink
                "card_bg": "#ffebf2",  # Lighter pink for cards
                "accent": "#ffacc7",  # Medium pink accent
                "accent_hover": "#ff85a1",  # Darker pink for hover
                "text_primary": "#4a4a4a",  # Dark gray for text
                "dropdown_bg": "#ffebf2",  # Lighter pink for dropdown
                "dropdown_hover": "#ffd6e0",  # Medium light pink for hover
                "input_bg": "#ffebf2",  # Lighter pink for input
                "border": "#ffacc7",  # Medium pink for borders
                "file_bg": "#ffd6e0",  # Medium light pink for file rows
                "file_hover": "#ffc1d5",  # Slightly darker pink for file row hover
                "tree_bg": "#fff5f9",  # Very light pink for tree
                "tree_even": "#fff0f5",  # Light pink
                "tree_odd": "#ffebf2",  # Lighter pink
                "tree_header": "#ffacc7"  # Medium pink for headers
            }

    def create_copy(self):
        year = self.year_entry.get().strip()
        if not year or not self._validate_year(year):
            self.status_label.configure(text="Please enter a valid year range (e.g. 2024-2025)", text_color="red")
            return
            
        # Get selected term - exact sheet name as it appears in Excel
        term = self.term_var.get()  # Now directly "term1" or "term2"
        
        # Get selected standard
        std = self.std_var.get()  # "FYJC" or "SYJC"
        
        os.makedirs("excel_copies", exist_ok=True)
        new_file = f"excel_copies/iso_excel_{year}_{term}_{std}.xlsx"
        
        try:
            import openpyxl
            
            # Open the source file
            src_wb = openpyxl.load_workbook("iso_excel.xlsx")
            
            # Check if the term sheet exists
            if term not in src_wb.sheetnames:
                self.status_label.configure(text=f"Error: Sheet '{term}' not found in template", text_color="red")
                return
                
            # Create a new workbook
            dst_wb = openpyxl.Workbook()
            
            # Get the source sheet by name (exactly as it appears in Excel)
            src_sheet = src_wb[term]
            
            # Get destination default sheet
            dst_sheet = dst_wb.active
            dst_sheet.title = term
            
            # Copy cell values, styles, merged cells, etc.
            for row in src_sheet.rows:
                for cell in row:
                    dst_cell = dst_sheet.cell(row=cell.row, column=cell.column)
                    dst_cell.value = cell.value
                    if cell.has_style:
                        dst_cell.font = copy(cell.font)
                        dst_cell.border = copy(cell.border)
                        dst_cell.fill = copy(cell.fill)
                        dst_cell.number_format = cell.number_format
                        dst_cell.protection = copy(cell.protection)
                        dst_cell.alignment = copy(cell.alignment)
            
            # Copy column dimensions
            for col, width in src_sheet.column_dimensions.items():
                dst_sheet.column_dimensions[col].width = width.width
                
            # Copy row dimensions
            for row, height in src_sheet.row_dimensions.items():
                dst_sheet.row_dimensions[row].height = height.height
                
            # Copy merged cells
            for merged_cell_range in src_sheet.merged_cells.ranges:
                dst_sheet.merge_cells(str(merged_cell_range))
            
            # Add standard information at the bottom of the sheet
            # First, find the last row with data
            last_row = dst_sheet.max_row
            
            # Add a small gap (1 row)
            std_row = last_row + 2
            
            # Add standard info with proper styling
            std_info_cell = dst_sheet.cell(row=std_row, column=1)
            std_info_cell.value = f"Standard: {std} ({self._std_label_map(std)})"
            
            # Style the standard information cell
            std_info_cell.font = openpyxl.styles.Font(name='Arial', size=10, bold=True)
            std_info_cell.alignment = openpyxl.styles.Alignment(horizontal='left')
            
            # Create a border for the standard information
            thin_border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
            std_info_cell.border = thin_border
            
            # Add a light fill color
            if std == "FYJC":
                # Light purple for FYJC
                std_info_cell.fill = openpyxl.styles.PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            else:
                # Light pink for SYJC
                std_info_cell.fill = openpyxl.styles.PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            # Save the workbook
            dst_wb.save(new_file)
            
            # Display success with cute emojis
            self.status_label.configure(text=f"✅ Copy created for {std}: {new_file}", text_color="green")
            self.new_file_path = os.path.abspath(new_file)
            self.refresh_file_list()
            
        except Exception as e:
            self.status_label.configure(text=f"❌ Error: {e}", text_color="red")
    
    def _std_label_map(self, std):
        """Convert standard code to descriptive text"""
        if std == "FYJC":
            return "11th Standard"
        elif std == "SYJC":
            return "12th Standard"
        return ""

    def refresh_file_list(self):
        """Refresh the list of available Excel files, filtering out system files."""
        # Update colors in case appearance mode changed
        self.update_colors()
        
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        
        # Get all files first (without filtering)
        self.all_files = []
        available_years = set(["All Years"])
        
        if os.path.exists("excel_copies"):
            # Get only valid Excel files, filtering out system files and temp files
            all_files = os.listdir("excel_copies")
            excel_files = [f for f in all_files if (
                f.endswith(".xlsx") and  # Only Excel files
                not f.startswith("~") and  # Not temp files
                not f.startswith("$") and  # Not system files
                not f.startswith(".")  # Not hidden files
            )]
            
            # Extract metadata from filenames for filtering
            for fname in excel_files:
                file_info = self._parse_filename(fname)
                self.all_files.append((fname, file_info))
                
                # Collect available years for the filter
                if file_info["year"]:
                    available_years.add(file_info["year"])
        
        # Update year filter options
        self.year_filter.configure(values=sorted(list(available_years)))
        
        # Apply current filters
        self._display_filtered_files()
    
    def _parse_filename(self, filename):
        """Extract year, term and standard information from filename"""
        info = {
            "year": "",
            "term": "",
            "std": ""
        }
        
        # Example filename format: iso_excel_2024-2025_term1_FYJC.xlsx
        parts = filename.replace(".xlsx", "").split("_")
        
        # Extract year
        for part in parts:
            if "-" in part and part.startswith("20"):
                info["year"] = part
                break
        
        # Extract term - look for exact term matching
        if "_term1" in filename:
            info["term"] = "term1"
        elif "_term2" in filename:
            info["term"] = "term2"
        
        # Extract standard - look for exact std matching
        if "_FYJC" in filename:
            info["std"] = "FYJC"
        elif "_SYJC" in filename:
            info["std"] = "SYJC"
        
        return info
    
    def _display_filtered_files(self):
        """Display files based on current filter settings"""
        # Get current filter values
        year_filter = self.year_filter_var.get()
        term_filter = self.term_filter_var.get()
        std_filter = self.std_filter_var.get()
        
        # Apply filters
        filtered_files = []
        for fname, info in self.all_files:
            # Check if the file matches all active filters
            if (year_filter == "All Years" or info["year"] == year_filter) and \
               (term_filter == "All Terms" or info["term"] == term_filter) and \
               (std_filter == "All Standards" or info["std"] == std_filter):
                filtered_files.append((fname, info))
        
        # Display the filtered files
        if not filtered_files:
            no_files_frame = ctk.CTkFrame(self.scrollable_frame, fg_color="transparent")
            no_files_frame.grid(row=0, column=0, sticky="ew", pady=20, padx=20)
            
            # Add a cute empty state message with emoji
            ctk.CTkLabel(no_files_frame, 
                        text="📭", 
                        font=ctk.CTkFont(size=40)).grid(row=0, column=0, pady=(20, 10))
            
            filter_message = "matching your filters" if year_filter != "All Years" or term_filter != "All Terms" or std_filter != "All Standards" else "found"
            ctk.CTkLabel(no_files_frame, 
                        text=f"No copies {filter_message}.", 
                        font=ctk.CTkFont(size=18, weight="bold"),
                        text_color=self.colors["accent"]).grid(row=1, column=0, pady=(10, 20))
        else:
            # Configure the scrollable frame
            self.scrollable_frame.grid_columnconfigure(0, weight=1)
            
            # Create a parent frame for all file items to ensure proper layout
            files_container = ctk.CTkFrame(self.scrollable_frame, fg_color="transparent")
            files_container.grid(row=0, column=0, sticky="ew", padx=5, pady=5)
            files_container.grid_columnconfigure(0, weight=1)
            
            for i, (fname, info) in enumerate(sorted(filtered_files), start=0):
                file_path = os.path.abspath(os.path.join("excel_copies", fname))
                
                # Create a compact file row with hover effect
                row_frame = ctk.CTkFrame(files_container, 
                                       fg_color=self.colors["file_bg"], 
                                       corner_radius=10)
                row_frame.grid(row=i, column=0, sticky="ew", pady=3, padx=5)  # Reduced padding
                
                # Configure row layout
                row_frame.grid_columnconfigure(0, weight=0)  # File icon
                row_frame.grid_columnconfigure(1, weight=0)  # Filename
                row_frame.grid_columnconfigure(2, weight=1)  # Badges (expand to push button right)
                row_frame.grid_columnconfigure(3, weight=0)  # Button fixed size
                
                # File icon - more compact
                file_icon = ctk.CTkLabel(row_frame, 
                                      text="📄", 
                                      font=ctk.CTkFont(size=18))
                file_icon.grid(row=0, column=0, padx=(8, 5), pady=6, sticky="w")
                
                # Filename - more compact
                short_name = self._shorten_filename(fname)
                file_name = ctk.CTkLabel(row_frame, 
                                       text=short_name, 
                                       font=ctk.CTkFont(size=14, weight="bold"),
                                       text_color=self.colors["text_primary"])
                file_name.grid(row=0, column=1, padx=(0, 8), pady=6, sticky="w")
                
                # Badges container - compact horizontal layout
                badges_frame = ctk.CTkFrame(row_frame, fg_color="transparent")
                badges_frame.grid(row=0, column=2, padx=0, pady=6, sticky="w")
                
                # Add badges in a row
                badge_idx = 0
                
                if info["year"]:
                    year_badge = self._create_badge(badges_frame, f"{info['year']}", "#3a2b4a")
                    year_badge.grid(row=0, column=badge_idx, padx=(0, 3))
                    badge_idx += 1
                
                if info["term"]:
                    term_text = info["term"].replace("term", "T")  # Shorter text
                    term_badge = self._create_badge(badges_frame, term_text, "#473960")
                    term_badge.grid(row=0, column=badge_idx, padx=(0, 3))
                    badge_idx += 1
                
                if info["std"]:
                    std_bg = "#E6E6FA" if info["std"] == "FYJC" else "#FFE6E6"
                    std_badge = self._create_badge(badges_frame, info["std"], std_bg, text_color="#333333")
                    std_badge.grid(row=0, column=badge_idx, padx=(0, 3))
                
                # Open button - more compact
                open_btn = ctk.CTkButton(row_frame, 
                                       text="Open", 
                                       width=60,  # Smaller button
                                       height=24,
                                       font=ctk.CTkFont(size=12, weight="bold"),
                                       fg_color=self.colors["accent"],
                                       hover_color=self.colors["accent_hover"],
                                       corner_radius=8,
                                       command=lambda p=file_path: webbrowser.open(f"file://{p}"))
                open_btn.grid(row=0, column=3, padx=(5, 8), pady=6, sticky="e")
                
                # Create highlight effect on hover
                def on_enter(e, frame=row_frame):
                    frame.configure(fg_color=self.colors["file_hover"])
                    
                def on_leave(e, frame=row_frame):
                    frame.configure(fg_color=self.colors["file_bg"])
                    
                row_frame.bind("<Enter>", on_enter)
                row_frame.bind("<Leave>", on_leave)
    
    def _shorten_filename(self, filename):
        """Create a shorter display version of the filename"""
        # Remove the common prefix
        if filename.startswith("iso_excel_"):
            filename = filename[10:]
        # Remove .xlsx extension
        if filename.endswith(".xlsx"):
            filename = filename[:-5]
        return filename
    
    def _create_badge(self, parent, text, bg_color, text_color="#ffffff"):
        """Create a small badge with metadata"""
        badge = ctk.CTkFrame(parent, fg_color=bg_color, corner_radius=5)  # Smaller radius
        
        ctk.CTkLabel(
            badge,
            text=text,
            font=ctk.CTkFont(size=10),  # Smaller font
            text_color=text_color,
            padx=4,  # Reduced padding
            pady=0
        ).pack(padx=2, pady=1)  # Reduced padding
        
        return badge
    
    def apply_filters(self, value=None):
        """Apply filters to the file list"""
        # Update display based on current filter values
        self._display_filtered_files()
    
    def reset_filters(self):
        """Reset all filters to default values"""
        self.year_filter_var.set("All Years")
        self.term_filter_var.set("All Terms")
        self.std_filter_var.set("All Standards")
        # Apply the reset filters
        self._display_filtered_files()

    def _validate_year(self, year):
        import re
        return re.match(r"^20\d{2}-20\d{2}$", year)

class SplashScreen(ctk.CTkToplevel):
    def __init__(self, app):
        super().__init__()
        self.app = app
        self.alpha = 0.0  # Start fully transparent
        
        # Set up the splash window
        self.overrideredirect(True)  # No window decorations
        self.wm_attributes("-topmost", True)  # Keep on top
        self.attributes("-alpha", self.alpha)  # Set initial transparency
        
        # Calculate position (center on screen)
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        splash_width = 650
        splash_height = 450
        x = (screen_width - splash_width) // 2
        y = (screen_height - splash_height) // 2
        self.geometry(f"{splash_width}x{splash_height}+{x}+{y}")
        
        # Create a frame with a cute pastel background
        if ctk.get_appearance_mode() == "Dark":
            bg_color = "#2d2438"  # Dark purple
            inner_color = "#3a2b4a"  # Medium purple
            border_color = "#b76edc"  # Bright purple
            title_color = "#e2b6ff"  # Light purple
            subtitle_color = "#c78ae8"  # Medium light purple
        else:
            bg_color = "#fff0f5"  # Light pink
            inner_color = "#ffebf2"  # Lighter pink
            border_color = "#ffacc7"  # Medium pink
            title_color = "#ff85a1"  # Darker pink
            subtitle_color = "#ffacc7"  # Medium pink
            
        self.configure(fg_color=bg_color)
        
        # Main content frame with cute border
        content_frame = ctk.CTkFrame(self, fg_color=inner_color, corner_radius=25, 
                                   border_width=4, border_color=border_color)
        content_frame.grid(row=0, column=0, sticky="nsew", padx=40, pady=40)
        content_frame.grid_columnconfigure(0, weight=1)
        content_frame.grid_rowconfigure(3, weight=0)
        
        # Grid configuration for the main window
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        
        # Top decoration - row of cute emojis
        top_emojis = ctk.CTkFrame(content_frame, fg_color="transparent")
        top_emojis.grid(row=0, column=0, pady=(20, 0))
        
        for i, emoji in enumerate(["🌸", "✨", "💖", "✨", "🌸"]):
            ctk.CTkLabel(
                top_emojis,
                text=emoji,
                font=ctk.CTkFont(size=30),
            ).grid(row=0, column=i, padx=15)
        
        # Welcome text with cute styling
        welcome_label = ctk.CTkLabel(
            content_frame,
            text="Welcome Mummy",
            font=ctk.CTkFont(family="Arial", size=42, weight="bold"),
            text_color=title_color
        )
        welcome_label.grid(row=1, column=0, pady=(20, 0))
        
        # Subtitle text
        subtitle_label = ctk.CTkLabel(
            content_frame,
            text="to your ISO Manager",
            font=ctk.CTkFont(family="Arial", size=24, weight="bold"),
            text_color=subtitle_color
        )
        subtitle_label.grid(row=2, column=0, pady=(5, 30))
        
        # Create a cute image frame
        image_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        image_frame.grid(row=3, column=0, sticky="ew", pady=(0, 20))
        
        # Add spreadsheet icon with big emoji
        spreadsheet_label = ctk.CTkLabel(
            image_frame,
            text="📊",
            font=ctk.CTkFont(size=80),
        )
        spreadsheet_label.pack(pady=5)
        
        # Prettier, gradient-like progress bar
        progress_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        progress_frame.grid(row=4, column=0, sticky="ew", pady=(10, 10), padx=80)
        progress_frame.grid_columnconfigure(0, weight=1)
        
        self.spinner = ctk.CTkProgressBar(progress_frame, width=400, height=20, 
                                        corner_radius=10, 
                                        progress_color=border_color,
                                        fg_color="#473960" if ctk.get_appearance_mode() == "Dark" else "#ffe0e9")
        self.spinner.grid(row=0, column=0, pady=10)
        self.spinner.set(0)
        
        # Loading text
        self.loading_label = ctk.CTkLabel(
            content_frame,
            text="Loading...",
            font=ctk.CTkFont(family="Arial", size=18, weight="bold"),
            text_color=subtitle_color
        )
        self.loading_label.grid(row=5, column=0, pady=(5, 20))
        
        # Bottom decoration - row of cute emojis
        bottom_emojis = ctk.CTkFrame(content_frame, fg_color="transparent")
        bottom_emojis.grid(row=6, column=0, pady=(0, 20))
        
        for i, emoji in enumerate(["💕", "✨", "🎀", "✨", "💕"]):
            ctk.CTkLabel(
                bottom_emojis,
                text=emoji,
                font=ctk.CTkFont(size=30),
            ).grid(row=0, column=i, padx=15)
        
        # Start animation sequence
        self.after(100, self.fade_in)
    
    def fade_in(self):
        """Fade in the splash screen"""
        if self.alpha < 1.0:
            self.alpha += 0.05
            self.attributes("-alpha", self.alpha)
            self.after(20, self.fade_in)
        else:
            # Once fully visible, start the progress animation
            self.after(100, lambda: self.animate_progress(0))
    
    def animate_progress(self, progress):
        """Animate the progress bar with cute loading messages"""
        loading_messages = [
            "Loading your data... 💾",
            "Preparing your workspace... 🎀",
            "Sprinkling some sparkles... ✨",
            "Almost ready... 🌈",
            "Just a moment... 💫",
            "Setting things up for you... 🌸"
        ]
        
        if progress <= 1.0:
            self.spinner.set(progress)
            
            # Update loading message periodically
            if progress % 0.15 < 0.02 and progress > 0:
                msg_index = int(min(progress * 6, 5))
                self.loading_label.configure(text=loading_messages[msg_index])
                
            self.after(30, lambda: self.animate_progress(progress + 0.01))
        else:
            # Show completion message
            self.loading_label.configure(text="Ready to go! 🎉")
            # After progress completes, fade out
            self.after(800, self.fade_out)
    
    def fade_out(self):
        """Fade out the splash screen"""
        if self.alpha > 0:
            self.alpha -= 0.05
            self.attributes("-alpha", self.alpha)
            self.after(20, self.fade_out)
        else:
            # After fade out, destroy splash and show main app
            self.destroy()
            self.app.deiconify()  # Show the main app window

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("ISO")
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        self.geometry(f"{screen_width}x{screen_height}+0+0")
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        
        # Set cute color scheme - dark mode only
        self.cute_colors = {
            "bg_primary": "#2d2438",  # Dark purple background
            "bg_secondary": "#3a2b4a",  # Medium purple
            "accent": "#b76edc",  # Bright purple accent
            "accent_hover": "#c78ae8",  # Lighter purple for hover
            "text_primary": "#ffffff",  # White text for better visibility
            "text_light": "#ffffff",  # White text
            "title_bg": "#b76edc",  # Purple title background
            "button_bg": "#3a2b4a",  # Dark purple button background
            "button_active": "#b76edc",  # Bright purple for active button
            "content_bg": "#2d2438"  # Dark purple content background
        }
        
        # Always use dark mode
        ctk.set_appearance_mode("dark")
        
        # Initially hide the main window
        self.withdraw()
        
        # Sidebar with cute styling
        self.sidebar = ctk.CTkFrame(self, width=240, fg_color=self.cute_colors["bg_primary"], corner_radius=0)
        self.sidebar.grid(row=0, column=0, sticky="nsw", padx=0, pady=0)
        self.sidebar.grid_propagate(False)
        self.sidebar.grid_rowconfigure(10, weight=1)
        self.sidebar.grid_columnconfigure(0, weight=1)
        
        # App title with cute styling - store reference for theme changing
        self.title_frame = ctk.CTkFrame(self.sidebar, fg_color=self.cute_colors["title_bg"], height=120, corner_radius=0)
        self.title_frame.grid(row=0, column=0, sticky="ew", padx=0, pady=(0, 20))
        self.title_frame.grid_propagate(False)
        self.title_frame.grid_columnconfigure(0, weight=1)
        
        # Add cute icon before title
        title_content = ctk.CTkFrame(self.title_frame, fg_color="transparent")
        title_content.grid(row=0, column=0, padx=10, pady=20)
        title_content.grid_columnconfigure(1, weight=1)
        
        # Left cute icon - larger
        ctk.CTkLabel(title_content, 
                     text="📊", 
                     font=ctk.CTkFont(size=40)).grid(row=0, column=0, padx=(5, 15))
        
        # Make title text bigger and ensure it's white for contrast
        self.title_label = ctk.CTkLabel(title_content, 
                                  text="ISO", 
                                  font=ctk.CTkFont(size=36, weight="bold", family="Arial"),
                                  text_color="#ffffff")
        self.title_label.grid(row=0, column=1, padx=5)
        
        # Right cute icon - larger
        ctk.CTkLabel(title_content, 
                     text="✨", 
                     font=ctk.CTkFont(size=40)).grid(row=0, column=2, padx=(15, 5))
        
        # Navigation buttons with cute styling
        button_frame = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        button_frame.grid(row=1, column=0, padx=15, pady=15, sticky="ew")
        
        self.copy_button = ctk.CTkButton(button_frame, 
                                        text="Manage Excel Copies", 
                                        command=self.show_copy, 
                                        font=ctk.CTkFont(size=16, weight="bold"),
                                        height=50,
                                        fg_color=self.cute_colors["button_bg"],
                                        text_color=self.cute_colors["text_primary"],
                                        hover_color=self.cute_colors["accent"],
                                        corner_radius=15)
        self.copy_button.grid(row=0, column=0, padx=10, pady=10, sticky="ew")
        
        self.edit_button = ctk.CTkButton(button_frame, 
                                        text="Edit Data", 
                                        command=self.show_edit, 
                                        font=ctk.CTkFont(size=16, weight="bold"),
                                        height=50,
                                        fg_color=self.cute_colors["button_bg"],
                                        text_color=self.cute_colors["text_primary"],
                                        hover_color=self.cute_colors["accent"],
                                        corner_radius=15)
        self.edit_button.grid(row=1, column=0, padx=10, pady=10, sticky="ew")
        
        # Decorative elements
        for i, emoji in enumerate(["🌸", "💜", "🌸"]):
            ctk.CTkLabel(self.sidebar, 
                         text=emoji, 
                         font=ctk.CTkFont(size=24)).grid(row=2+i, column=0, pady=5)
        
        # Main content with cute styling
        self.content_frame = ctk.CTkFrame(self, fg_color=self.cute_colors["content_bg"], corner_radius=20)
        self.content_frame.grid(row=0, column=1, sticky="nsew", padx=20, pady=20)
        self.content_frame.grid_columnconfigure(0, weight=1)
        self.content_frame.grid_rowconfigure(0, weight=1)
        
        self.pages = {
            "copy": CopyPage(self.content_frame),
            "edit": EditPage(self.content_frame)
        }
        self.current_page = None
        
        # Footer with cute design
        footer_frame = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        footer_frame.grid(row=15, column=0, sticky="ew", padx=15, pady=15)
        
        ctk.CTkLabel(footer_frame, 
                    text="Made with 💖",
                    font=ctk.CTkFont(size=14),
                    text_color=self.cute_colors["accent"]).pack(pady=5)
        
        # Create splash screen after initializing the main app
        self.splash = SplashScreen(self)
        
        # Show initial page
        self.show_copy()

    def show_copy(self):
        self._show_page("copy")
        self.copy_button.configure(fg_color=self.cute_colors["accent"])
        self.edit_button.configure(fg_color=self.cute_colors["button_bg"])
        self.pages["copy"].refresh_file_list()
        
    def show_edit(self):
        self._show_page("edit")
        self.edit_button.configure(fg_color=self.cute_colors["accent"])
        self.copy_button.configure(fg_color=self.cute_colors["button_bg"])
        self.pages["edit"].load_files()
        
    def _show_page(self, page):
        if self.current_page:
            self.current_page.grid_remove()
        self.pages[page].grid(row=0, column=0, sticky="nsew")
        self.current_page = self.pages[page]

if __name__ == "__main__":
    # Always use dark mode
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")
    
    # Set fonts
    try:
        # Try to use a cute font if available
        if "Arial" in tk.font.families():
            default_font = "Arial"
        elif "Helvetica" in tk.font.families():
            default_font = "Helvetica"
        else:
            default_font = None
            
        if default_font:
            default_font_size = 14
            tk.font.nametofont("TkDefaultFont").configure(family=default_font, size=default_font_size)
            tk.font.nametofont("TkTextFont").configure(family=default_font, size=default_font_size)
    except Exception:
        pass
    
    app = App()
    app.mainloop()
