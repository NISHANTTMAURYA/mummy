import customtkinter as ctk
import os
import webbrowser
import shutil
import openpyxl
import tkinter as tk
from tkinter import ttk
from copy import copy

class ExcelPage(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent)
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.open_button = ctk.CTkButton(
            self,
            text="Open Excel File",
            command=self.open_excel_file,
            font=ctk.CTkFont(size=16, weight="bold")
        )
        self.open_button.grid(row=0, column=0, padx=40, pady=40)

    def open_excel_file(self):
        excel_path = os.path.abspath("iso_excel.xlsx")
        webbrowser.open(f"file://{excel_path}")

class EditPage(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent)
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(4, weight=1)
        
        # Header frame with improved styling
        header_frame = ctk.CTkFrame(self, fg_color="transparent")
        header_frame.grid(row=0, column=0, columnspan=2, sticky="ew", padx=40, pady=(30, 20))
        header_frame.grid_columnconfigure(1, weight=1)
        
        # File selection with improved styling
        ctk.CTkLabel(header_frame, text="Select Excel File:", 
                    font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, sticky="w", pady=(0, 10))
        self.file_var = ctk.StringVar()
        self.file_menu = ctk.CTkOptionMenu(header_frame, variable=self.file_var, 
                                          values=self.get_file_list(), 
                                          command=self.on_file_change, 
                                          width=400,
                                          height=40,
                                          font=ctk.CTkFont(size=14))
        self.file_menu.grid(row=0, column=1, sticky="w", padx=(20, 0), pady=(0, 10))
        
        # Month selection with improved styling
        ctk.CTkLabel(header_frame, text="Select Month:", 
                    font=ctk.CTkFont(size=16, weight="bold")).grid(row=1, column=0, sticky="w", pady=(10, 0))
        self.month_var = ctk.StringVar()
        self.month_menu = ctk.CTkOptionMenu(header_frame, variable=self.month_var, 
                                           values=[], 
                                           command=self.on_month_change, 
                                           width=200,
                                           height=40,
                                           font=ctk.CTkFont(size=14))
        self.month_menu.grid(row=1, column=1, sticky="w", padx=(20, 0), pady=(10, 0))
        
        # Refresh button
        self.refresh_button = ctk.CTkButton(header_frame,
                                           text="Refresh",
                                           command=self.refresh_data,
                                           font=ctk.CTkFont(size=14, weight="bold"),
                                           width=100,
                                           height=40,
                                           fg_color="#1f538d",
                                           hover_color="#3a7ebf")
        self.refresh_button.grid(row=1, column=1, sticky="e", padx=(0, 20), pady=(10, 0))
        
        # Data frame for the table with improved styling
        self.data_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.data_frame.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=40, pady=20)
        self.data_widgets = []
        
        # Save button with improved styling
        button_frame = ctk.CTkFrame(self, fg_color="transparent")
        button_frame.grid(row=3, column=0, columnspan=2, sticky="ew", padx=40, pady=(10, 20))
        button_frame.grid_columnconfigure(0, weight=1)
        
        self.save_button = ctk.CTkButton(button_frame, 
                                        text="Save Changes", 
                                        command=self.save_changes, 
                                        font=ctk.CTkFont(size=16, weight="bold"),
                                        height=50,
                                        width=200,
                                        fg_color="#1f538d",
                                        hover_color="#3a7ebf")
        self.save_button.grid(row=0, column=0, pady=10)
        
        # Status label with improved styling
        self.status_label = ctk.CTkLabel(button_frame, 
                                        text="", 
                                        font=ctk.CTkFont(size=14))
        self.status_label.grid(row=1, column=0, pady=(5, 10))
        
        self.current_file = None
        self.current_month = None
        self.month_col_ranges = {}
        self.sub_headers = []
        self.initials = []
        self._last_edited_cell = None
        self._refresh_pending = False
        self.load_files()

    def get_file_list(self):
        """Get a filtered list of Excel files for the file selection dropdown."""
        if not os.path.exists("excel_copies"):
            return []
            
        # Get only valid Excel files, filtering out system files and temp files
        all_files = os.listdir("excel_copies")
        return [f for f in all_files if (
            f.endswith(".xlsx") and  # Only Excel files
            not f.startswith("~") and  # Not temp files
            not f.startswith("$") and  # Not system files
            not f.startswith(".")  # Not hidden files
        )]

    def load_files(self):
        files = self.get_file_list()
        self.file_menu.configure(values=files)
        if files:
            self.file_var.set(files[0])
            self.on_file_change(files[0])
        else:
            self.file_var.set("")
            self.month_menu.configure(values=[])
            self.clear_data_frame()

    def on_file_change(self, filename):
        import openpyxl
        self.current_file = os.path.join("excel_copies", filename)
        try:
            # When parsing the file structure, use normal mode to get formulas
            wb = openpyxl.load_workbook(self.current_file)
            ws = wb.active
            
            # First, find the months and their column ranges by analyzing the first row
            months = []
            month_col_ranges = {}
            
            # Get all cells from the first row
            first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            second_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
            
            # Scan columns to find month headers
            current_month = None
            start_col = None
            
            for idx, val in enumerate(first_row):
                if idx < 2:  # Skip SR.NO. and INITIALS columns
                    continue
                    
                # If we find a month or TOTAL header
                if val and isinstance(val, str) and val.strip() and val.strip().upper() not in ("SR.NO.", "INITIALS"):
                    # If we were tracking a previous month, save its range
                    if current_month:
                        month_col_ranges[current_month] = (start_col, idx-1)
                    
                    # Start tracking new month
                    current_month = str(val).strip()
                    start_col = idx
                    months.append(current_month)
            
            # Save the last month's range
            if current_month:
                month_col_ranges[current_month] = (start_col, len(first_row)-1)
            
            # Store the column ranges and headers
            self.month_col_ranges = month_col_ranges
            self.sub_headers = second_row  # Second row contains column headers like ALOTTED, E-Act, etc.
            
            # Get all initials (skip TOTAL/empty)
            initials = []
            for row in ws.iter_rows(min_row=3, max_col=2, values_only=True):
                if row[1] and str(row[1]).strip() and str(row[1]).strip().upper() != "TOTAL":
                    initials.append(str(row[1]).strip())
            
            self.initials = initials
            self.month_menu.configure(values=months)
            
            if months:
                self.month_var.set(months[0])
                self.on_month_change(months[0])
            else:
                self.month_var.set("")
                self.clear_data_frame()
                
        except Exception as e:
            self.status_label.configure(text=f"Error loading file: {e}", text_color="red")
            self.month_menu.configure(values=[])
            self.clear_data_frame()

    def on_month_change(self, month):
        """Handle month selection change"""
        self.current_month = month
        # Clear any existing edit widgets
        if hasattr(self, 'edit_entry') and self.edit_entry:
            edit_frame, _, _, _, _ = self.edit_entry
            if edit_frame.winfo_exists():
                edit_frame.destroy()
            self.edit_entry = None
        # Display data for the selected month
        self.display_data()

    def clear_data_frame(self):
        for widget in self.data_frame.winfo_children():
            widget.destroy()
        self.data_widgets = []
        self.tree = None
        self.tree_vars = {}

    def display_data(self):
        self.clear_data_frame()
        if not (self.current_file and self.current_month):
            return
            
        import openpyxl
        try:
            # Unbind previous resize handlers
            try:
                self.winfo_toplevel().unbind("<Configure>")
            except:
                pass
                
            # Load workbook with data_only=True to get calculated values instead of formulas
            wb = openpyxl.load_workbook(self.current_file, data_only=True)
            ws = wb.active
            
            # Get the column range for the selected month
            col_range = self.month_col_ranges.get(self.current_month)
            if not col_range:
                self.status_label.configure(text="Month not found.", text_color="red")
                ctk.CTkLabel(self.data_frame, text="No data found for selected month.", font=ctk.CTkFont(size=14)).grid(row=0, column=0, padx=10, pady=10)
                return
                
            start_col, end_col = col_range
            
            # Define which headers we want to display/edit
            editable_headers = ["ALOTTED", "E-Act", "E-Add"]
            headers = ["Initial"]
            header_indices = []
            
            # Find the indices of editable columns within the month's range
            for col in range(start_col, end_col+1):
                if col < len(self.sub_headers):  # Ensure we don't go out of bounds
                    header = self.sub_headers[col]
                    if header in editable_headers:
                        headers.append(header)
                        header_indices.append(col)
            
            # Gather all initials and their values for the selected month
            data = []
            row_indices = []
            
            for i, row in enumerate(ws.iter_rows(min_row=3, max_col=2, values_only=True), start=3):
                initial = row[1]
                if initial and str(initial).strip() and str(initial).strip().upper() != "TOTAL":
                    values = [initial]
                    
                    for col, h in zip(header_indices, headers[1:]):
                        # Ensure we don't try to access columns outside worksheet bounds
                        if col+1 <= ws.max_column:
                            cell = ws.cell(row=i, column=col+1)
                            val = cell.value
                            
                            if h == "E-Add":
                                if val == '+' or val == '+ ' or val is None:
                                    val = ''
                                elif val is not None:
                                    # Clean up the value - keep only the number part
                                    val = str(val).strip().replace('+', '').strip()
                            elif val is None or str(val).strip().lower() == 'none':
                                val = ''
                            else:
                                # Convert to string to display numbers properly
                                val = str(val)
                        else:
                            val = ''
                            
                        values.append(val)
                    
                    data.append(values)
                    row_indices.append(i)
                    
            if not headers or not data:
                ctk.CTkLabel(self.data_frame, text="No data to display for this selection.", font=ctk.CTkFont(size=14)).grid(row=0, column=0, padx=10, pady=10)
                return

            # Create a custom frame for the table with border
            table_container = ctk.CTkFrame(self.data_frame, fg_color="transparent")
            table_container.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
            table_container.grid_columnconfigure(0, weight=1)
            table_container.grid_rowconfigure(0, weight=1)
            
            # Add a border frame around the table with improved styling
            border_frame = ctk.CTkFrame(table_container, fg_color="#1e2433", border_width=3, border_color="#3a7ebf", corner_radius=10)
            border_frame.grid(row=0, column=0, sticky="nsew", padx=0, pady=0)
            border_frame.grid_columnconfigure(0, weight=1)
            border_frame.grid_rowconfigure(0, weight=1)
            
            # Configure style for the treeview with visible grid lines
            style = ttk.Style()
            style.theme_use("clam")  # Use clam theme which supports more customization
            
            # Configure the Treeview colors and font
            style.configure("Custom.Treeview", 
                background="#2d3748",  # Darker blue-gray background
                foreground="#e2e8f0",  # Light gray text for better contrast
                rowheight=50,  # Increased row height
                fieldbackground="#2d3748",
                borderwidth=2)  # Increased border width
            
            # Configure the header style with more prominent headers
            style.configure("Custom.Treeview.Heading",
                background="#2563eb",  # Brighter blue headers
                foreground="white",
                relief="raised",
                borderwidth=2,
                font=('Arial', 16, 'bold'))  # Increased font size and bold
            
            # Configure selection colors
            style.map('Custom.Treeview', 
                background=[('selected', '#3b82f6')],  # Bright blue selection
                foreground=[('selected', 'white')])
            
            # Configure the Treeview to show grid lines
            style.layout("Custom.Treeview", [
                ('Custom.Treeview.treearea', {'sticky': 'nswe'})
            ])
            style.configure("Custom.Treeview", 
                            borderwidth=0,
                            relief="flat")
                
            # Create Treeview with increased row height
            self.tree = ttk.Treeview(border_frame, columns=headers, show="headings", 
                                    height=len(data), style="Custom.Treeview")
            
            # Configure column widths and headings
            column_width = 180
            for h in headers:
                self.tree.heading(h, text=h)
                self.tree.column(h, width=column_width, anchor="center")
            
            # Add data to the treeview with alternating row colors
            for idx, row in enumerate(data):
                item_id = self.tree.insert("", "end", values=row)
                
                # Apply alternating row colors
                if idx % 2 == 1:
                    self.tree.item(item_id, tags=('odd_row',))
                else:
                    self.tree.item(item_id, tags=('even_row',))
            
            # Configure row styles with subtle alternating colors
            self.tree.tag_configure('odd_row', background='#374151', font=('Arial', 14))  # Dark blue-gray
            self.tree.tag_configure('even_row', background='#2d3748', font=('Arial', 14))  # Darker blue-gray
            
            # Add horizontal grid lines tag
            self.tree.tag_configure('bottom_line', background='#4b5563')
            
            # Add a custom scrollbar with improved styling
            scrollbar_style = ttk.Style()
            scrollbar_style.configure("Custom.Vertical.TScrollbar", 
                                      background="#3a7ebf", 
                                      troughcolor="#1e2433",
                                      bordercolor="#3a7ebf",
                                      arrowcolor="white")
            
            scrollbar = ttk.Scrollbar(border_frame, orient="vertical", 
                                      command=self.tree.yview,
                                      style="Custom.Vertical.TScrollbar")
            self.tree.configure(yscrollcommand=scrollbar.set)
            
            # Place the treeview and scrollbar
            self.tree.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
            scrollbar.grid(row=0, column=1, sticky="ns", pady=5)
            
            # Add a title for the table
            month_title = ctk.CTkLabel(
                self.data_frame, 
                text=f"{self.current_month} Data", 
                font=ctk.CTkFont(size=20, weight="bold"),
                text_color="#3b82f6"
            )
            month_title.grid(row=0, column=0, sticky="n", pady=(0, 0))
            
            self.data_frame.grid_columnconfigure(0, weight=1)
            self.data_frame.grid_rowconfigure(0, weight=1)
            
            # Create a custom entry widget for editing cells
            self.edit_entry = None
            
            # Bind the double-click event using a dedicated method
            self.tree.bind("<Double-1>", self.on_double_click)
            
            # Store for saving
            self.data_widgets = [(headers, row_indices, header_indices)]
            self.status_label.configure(text="", text_color="green")
        except Exception as e:
            self.status_label.configure(text=f"Error: {e}", text_color="red")
            ctk.CTkLabel(self.data_frame, text=f"Error: {e}", font=ctk.CTkFont(size=14)).grid(row=0, column=0, padx=10, pady=10)
            
    def _save_single_cell(self, item_id, col_idx, header, value):
        """Save a single cell value directly to Excel without reloading the entire sheet"""
        try:
            # Get the row index from the data_widgets
            if not self.data_widgets:
                return False
                
            headers, row_indices, header_indices = self.data_widgets[0]
            
            # Get the item index in the treeview
            item_index = self.tree.index(item_id)
            if item_index >= len(row_indices):
                return False
                
            # Get the Excel row and column
            row_idx = row_indices[item_index]
            col = header_indices[col_idx - 1]  # -1 because col_idx includes Initial column
            
            # Open the workbook (load with formulas for saving)
            import openpyxl
            wb = openpyxl.load_workbook(self.current_file)
            ws = wb.active
            
            # Process the value based on the header
            if header == "E-Add":
                # Process E-Add value
                v = str(value).strip() if value is not None else ""
                v = v.replace('"', '').replace("'", '').replace('+', '').strip()
                
                # Always save with a '+' prefix for E-Add values
                if v:
                    cell_value = f'+{v}'
                else:
                    cell_value = '+'
                
                # Get the actual cell
                cell = ws.cell(row=row_idx, column=col+1)
                # Set its value
                cell.value = cell_value
            else:
                # For other headers, just use the value as is
                cell_value = value if value and value.strip() else None
                ws.cell(row=row_idx, column=col+1).value = cell_value
            
            # Save the workbook
            wb.save(self.current_file)
            
            # Don't refresh when moving to next cell during rapid edits
            # We'll set a flag so we know this cell was just edited
            self._last_edited_cell = (item_id, col_idx)
            
            # Update status
            self.status_label.configure(text="Changes saved!", text_color="green")
            return True
        except Exception as e:
            self.status_label.configure(text=f"Error saving: {e}", text_color="red")
            return False

    def save_changes(self):
        """Save all changes to the Excel file."""
        if not self.current_file or not self.current_file.endswith(".xlsx"):
            self.status_label.configure(text="No file selected or not an Excel file.", text_color="red")
            return
        
        # First, check if we are currently editing a cell and save it if needed
        if hasattr(self, 'edit_entry') and self.edit_entry:
            edit_frame, entry, item, col_idx, header = self.edit_entry
            if edit_frame.winfo_exists():
                # Get the current value from the entry
                new_value = entry.get().strip()
                
                # Update the treeview
                values = list(self.tree.item(item, 'values'))
                old_value = values[col_idx]
                values[col_idx] = new_value
                self.tree.item(item, values=values)
                
                # Clean up
                edit_frame.destroy()
                self.edit_entry = None
                
                # If the value has changed, save it
                if old_value != new_value:
                    self._save_single_cell(item, col_idx, header, new_value)
        
        import openpyxl
        try:
            # Get the current values directly from the treeview
            all_items = self.tree.get_children()
            current_values = []
            for item in all_items:
                current_values.append(self.tree.item(item, 'values'))
            
            # Open the workbook
            wb = openpyxl.load_workbook(self.current_file)
            ws = wb.active
            
            # Get row indices from data_widgets
            if not self.data_widgets:
                self.status_label.configure(text="No data to save.", text_color="red")
                return
                
            headers, row_indices, header_indices = self.data_widgets[0]
            
            # Save each value
            for row_idx, values in zip(row_indices, current_values):
                for i, col in enumerate(header_indices):
                    header = headers[i+1]  # +1 because headers[0] is "Initial"
                    val = values[i+1]      # +1 because values[0] is Initial
                    
                    if header == "E-Add":
                        # Process E-Add value
                        v = str(val).strip() if val is not None else ""
                        v = v.replace('"', '').replace("'", '').replace('+', '').strip()
                        
                        # Always save with a '+' prefix for E-Add values
                        if v:
                            cell_value = f'+{v}'
                        else:
                            cell_value = '+'
                        
                        # Get the actual cell
                        cell = ws.cell(row=row_idx, column=col+1)
                        # Set its value
                        cell.value = cell_value
                    else:
                        # Process other values
                        if val is not None and str(val).strip() != '' and str(val).strip().lower() != 'none':
                            ws.cell(row=row_idx, column=col+1).value = val
                        else:
                            ws.cell(row=row_idx, column=col+1).value = None
            
            # Save the workbook
            wb.save(self.current_file)
            
            # Force reload of data after save all changes
            self.refresh_data()
            
            self.status_label.configure(text="All changes saved!", text_color="green")
            
        except Exception as e:
            self.status_label.configure(text=f"Error saving: {e}", text_color="red")

    def on_double_click(self, event):
        """Handle double-click on table cell"""
        if not hasattr(self, 'tree') or not self.tree:
            return
            
        item = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        if not item or not column:
            return
        
        # Get column index (skip if it's the Initial column)
        col_idx = int(column.replace('#','')) - 1
        if col_idx == 0:
            return
        
        # Get current value and position for edit
        x, y, width, height = self.tree.bbox(item, column)
        current_value = self.tree.item(item, 'values')[col_idx]
        
        # Get headers from data_widgets
        if not self.data_widgets:
            return
        headers = self.data_widgets[0][0]
        
        # For E-Add column, remove the '+' if present
        header = headers[col_idx]
        if header == "E-Add" and current_value and str(current_value).startswith('+'):
            current_value = str(current_value).replace('+', '').strip()
        
        # Create a frame for better control with improved styling
        edit_frame = tk.Frame(self.tree, bg="#3b82f6", highlightthickness=2, highlightbackground="#60a5fa")
        edit_frame.place(x=x, y=y, width=width, height=height)
        
        # Create the entry widget with larger font and improved styling
        entry_var = tk.StringVar(value=current_value if current_value else "")
        entry = tk.Entry(edit_frame, textvariable=entry_var, 
                         font=('Arial', 16, 'bold'),  # Increased font size and bold
                         bg="#3b82f6",
                         fg="white",
                         bd=0,
                         highlightthickness=0,
                         justify=tk.CENTER)  # Center-aligned text
        entry.pack(fill="both", expand=True)
        entry.focus_set()
        entry.select_range(0, tk.END)
        
        # Store reference to edit widgets
        self.edit_entry = (edit_frame, entry, item, col_idx, header)
        
        def save_edit(event=None, move_to_next=True):
            # Get the new value
            new_value = entry_var.get().strip()
            
            # Update the treeview
            values = list(self.tree.item(item, 'values'))
            old_value = values[col_idx]
            values[col_idx] = new_value
            self.tree.item(item, values=values)
            
            # Clean up
            edit_frame.destroy()
            self.edit_entry = None
            
            # If the value hasn't changed, don't save
            save_success = True
            if old_value != new_value:
                # Save changes to Excel directly
                save_success = self._save_single_cell(item, col_idx, header, new_value)
            
            # Move to the next cell if requested
            if move_to_next and event and event.keysym == 'Return' and save_success:
                self._move_to_next_cell(item, col_idx)
        
        def cancel_edit(event=None):
            edit_frame.destroy()
            self.edit_entry = None
        
        # Bind events
        entry.bind("<Return>", save_edit)
        entry.bind("<Escape>", cancel_edit)
        entry.bind("<FocusOut>", lambda e: save_edit(e, move_to_next=False))
    
    def _move_to_next_cell(self, current_item, current_col_idx):
        """Move to the next editable cell after editing"""
        if not self.tree or not hasattr(self, 'tree'):
            return
            
        if not self.data_widgets:
            return
            
        headers = self.data_widgets[0][0]
        num_columns = len(headers)
        
        # If we're not at the last column, move to the next column in the same row
        if current_col_idx < num_columns - 1:
            next_col_idx = current_col_idx + 1
            # Skip the Initial column (0)
            if next_col_idx == 0:
                next_col_idx = 1
                
            # Create column identifier
            next_column = f"#{next_col_idx+1}"  # +1 because treeview columns are 1-indexed
            
            # Simulate a double-click on the next cell
            bbox = self.tree.bbox(current_item, next_column)
            if bbox:
                x, y, _, _ = bbox
                event = type('Event', (), {'x': x+5, 'y': y+5})  # +5 to ensure we're inside the cell
                self.on_double_click(event)
        else:
            # We're at the last column, move to the first column of the next row
            items = self.tree.get_children()
            try:
                current_index = items.index(current_item)
                if current_index < len(items) - 1:
                    next_item = items[current_index + 1]
                    # Move to the first editable column (skip Initial)
                    next_column = "#2"  # Column index 1 (2nd column) is the first editable one
                    
                    # Simulate a double-click on the next cell
                    bbox = self.tree.bbox(next_item, next_column)
                    if bbox:
                        x, y, _, _ = bbox
                        event = type('Event', (), {'x': x+5, 'y': y+5})  # +5 to ensure we're inside the cell
                        self.on_double_click(event)
            except ValueError:
                pass  # Item not found in the list

    def refresh_data(self):
        """Refresh the current data view"""
        if self.current_month:
            self.display_data()
            self.status_label.configure(text="Data refreshed!", text_color="green")

class CopyPage(ctk.CTkFrame):
    def __init__(self, parent):
        super().__init__(parent)
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        # --- Top: Create Copy Section ---
        top_frame = ctk.CTkFrame(self, fg_color="transparent")
        top_frame.grid(row=0, column=0, sticky="ew", padx=40, pady=(40, 20))
        top_frame.grid_columnconfigure(0, weight=1)

        title = ctk.CTkLabel(top_frame, text="Create Yearly Excel Copy", 
                           font=ctk.CTkFont(size=24, weight="bold"))
        title.grid(row=0, column=0, sticky="w", pady=(0, 20))

        entry_frame = ctk.CTkFrame(top_frame, fg_color="#2b2b2b", corner_radius=10)
        entry_frame.grid(row=1, column=0, sticky="ew", pady=(10, 0))
        entry_frame.grid_columnconfigure(1, weight=1)

        self.year_entry = ctk.CTkEntry(entry_frame, 
                                     placeholder_text="Enter year range (e.g. 2024-2025)", 
                                     font=ctk.CTkFont(size=16), 
                                     width=300,
                                     height=40)
        self.year_entry.grid(row=0, column=0, padx=20, pady=20, sticky="w")
        
        # Add term selection dropdown
        term_frame = ctk.CTkFrame(entry_frame, fg_color="transparent")
        term_frame.grid(row=1, column=0, padx=20, pady=(0, 20), sticky="w")
        
        ctk.CTkLabel(term_frame, 
                   text="Select Term:", 
                   font=ctk.CTkFont(size=16, weight="bold")).grid(row=0, column=0, sticky="w", padx=(0, 10))
        
        self.term_var = ctk.StringVar(value="term1")
        self.term_dropdown = ctk.CTkOptionMenu(term_frame,
                                             values=["term1", "term2"],
                                             variable=self.term_var,
                                             width=150,
                                             height=40,
                                             font=ctk.CTkFont(size=14),
                                             dropdown_font=ctk.CTkFont(size=14))
        self.term_dropdown.grid(row=0, column=1, sticky="w")
        
        self.copy_button = ctk.CTkButton(entry_frame, 
                                       text="Create Copy", 
                                       command=self.create_copy, 
                                       font=ctk.CTkFont(size=16, weight="bold"), 
                                       width=150,
                                       height=40,
                                       fg_color="#1f538d",
                                       hover_color="#3a7ebf")
        self.copy_button.grid(row=0, column=1, padx=20, pady=20, sticky="e")
        
        self.status_label = ctk.CTkLabel(entry_frame, 
                                       text="", 
                                       font=ctk.CTkFont(size=14))
        self.status_label.grid(row=2, column=0, columnspan=2, sticky="w", padx=20, pady=(0, 20))

        # --- Bottom: Available Copies List ---
        list_frame = ctk.CTkFrame(self, fg_color="transparent")
        list_frame.grid(row=1, column=0, sticky="nsew", padx=40, pady=(20, 40))
        list_frame.grid_columnconfigure(0, weight=1)
        list_frame.grid_rowconfigure(1, weight=1)

        list_title = ctk.CTkLabel(list_frame, 
                                text="Available Copies:", 
                                font=ctk.CTkFont(size=20, weight="bold"))
        list_title.grid(row=0, column=0, sticky="w", pady=(0, 15))

        # Scrollable frame for file list with improved styling
        self.scrollable_frame = ctk.CTkScrollableFrame(list_frame, 
                                                     fg_color="#2b2b2b", 
                                                     corner_radius=10,
                                                     height=350)
        self.scrollable_frame.grid(row=1, column=0, sticky="nsew")
        self.scrollable_frame.grid_columnconfigure(0, weight=1)
        self.refresh_file_list()
        self.new_file_path = None

    def create_copy(self):
        year = self.year_entry.get().strip()
        if not year or not self._validate_year(year):
            self.status_label.configure(text="Please enter a valid year range (e.g. 2024-2025)", text_color="red")
            return
            
        # Get selected term - exact sheet name as it appears in Excel
        term = self.term_var.get()  # Now directly "term1" or "term2"
        
        os.makedirs("excel_copies", exist_ok=True)
        new_file = f"excel_copies/iso_excel_{year}_{term}.xlsx"
        
        try:
            import openpyxl
            
            # Open the source file
            src_wb = openpyxl.load_workbook("iso_excel.xlsx")
            
            # Check if the term sheet exists
            if term not in src_wb.sheetnames:
                self.status_label.configure(text=f"Error: Sheet '{term}' not found in template", text_color="red")
                return
                
            # Create a new workbook
            dst_wb = openpyxl.Workbook()
            
            # Get the source sheet by name (exactly as it appears in Excel)
            src_sheet = src_wb[term]
            
            # Get destination default sheet
            dst_sheet = dst_wb.active
            dst_sheet.title = term
            
            # Copy cell values, styles, merged cells, etc.
            for row in src_sheet.rows:
                for cell in row:
                    dst_cell = dst_sheet.cell(row=cell.row, column=cell.column)
                    dst_cell.value = cell.value
                    if cell.has_style:
                        dst_cell.font = copy(cell.font)
                        dst_cell.border = copy(cell.border)
                        dst_cell.fill = copy(cell.fill)
                        dst_cell.number_format = cell.number_format
                        dst_cell.protection = copy(cell.protection)
                        dst_cell.alignment = copy(cell.alignment)
            
            # Copy column dimensions
            for col, width in src_sheet.column_dimensions.items():
                dst_sheet.column_dimensions[col].width = width.width
                
            # Copy row dimensions
            for row, height in src_sheet.row_dimensions.items():
                dst_sheet.row_dimensions[row].height = height.height
                
            # Copy merged cells
            for merged_cell_range in src_sheet.merged_cells.ranges:
                dst_sheet.merge_cells(str(merged_cell_range))
            
            # Save the workbook
            dst_wb.save(new_file)
            
            self.status_label.configure(text=f"Copy created with sheet '{term}': {new_file}", text_color="#00ff00")
            self.new_file_path = os.path.abspath(new_file)
            self.refresh_file_list()
            
        except Exception as e:
            self.status_label.configure(text=f"Error: {e}", text_color="red")

    def refresh_file_list(self):
        """Refresh the list of available Excel files, filtering out system files."""
        for widget in self.scrollable_frame.winfo_children():
            widget.destroy()
        files = []
        
        if os.path.exists("excel_copies"):
            # Get only valid Excel files, filtering out system files and temp files
            all_files = os.listdir("excel_copies")
            files = [f for f in all_files if (
                f.endswith(".xlsx") and  # Only Excel files
                not f.startswith("~") and  # Not temp files
                not f.startswith("$") and  # Not system files
                not f.startswith(".")  # Not hidden files
            )]
            
        if not files:
            ctk.CTkLabel(self.scrollable_frame, 
                        text="No copies found.", 
                        font=ctk.CTkFont(size=16)).grid(row=0, column=0, sticky="w", pady=10, padx=10)
        else:
            for i, fname in enumerate(sorted(files), start=0):
                file_path = os.path.abspath(os.path.join("excel_copies", fname))
                row_frame = ctk.CTkFrame(self.scrollable_frame, fg_color="#333333", corner_radius=5)
                row_frame.grid(row=i, column=0, sticky="ew", pady=6, padx=10)
                row_frame.grid_columnconfigure(0, weight=1)
                
                file_label = ctk.CTkLabel(row_frame, 
                                        text=fname, 
                                        font=ctk.CTkFont(size=16))
                file_label.grid(row=0, column=0, sticky="w", padx=15, pady=12)
                
                open_btn = ctk.CTkButton(row_frame, 
                                       text="Open", 
                                       width=100,
                                       height=32,
                                       font=ctk.CTkFont(size=14),
                                       fg_color="#1f538d",
                                       hover_color="#3a7ebf",
                                       command=lambda p=file_path: webbrowser.open(f"file://{p}"))
                open_btn.grid(row=0, column=1, padx=(12, 15), pady=12)

    def _validate_year(self, year):
        import re
        return re.match(r"^20\d{2}-20\d{2}$", year)

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Excel File Manager")
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        self.geometry(f"{screen_width}x{screen_height}+0+0")
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        
        # Sidebar with improved styling
        self.sidebar = ctk.CTkFrame(self, width=240, fg_color="#1a1a1a")
        self.sidebar.grid(row=0, column=0, sticky="nsw", padx=0, pady=0)
        self.sidebar.grid_propagate(False)
        self.sidebar.grid_rowconfigure(10, weight=1)
        self.sidebar.grid_columnconfigure(0, weight=1)
        
        # App title
        title_frame = ctk.CTkFrame(self.sidebar, fg_color="#1f538d", height=80)
        title_frame.grid(row=0, column=0, sticky="ew", padx=0, pady=(0, 20))
        title_frame.grid_propagate(False)
        title_frame.grid_columnconfigure(0, weight=1)
        
        title_label = ctk.CTkLabel(title_frame, 
                                  text="Excel Manager", 
                                  font=ctk.CTkFont(size=22, weight="bold"),
                                  text_color="white")
        title_label.grid(row=0, column=0, padx=20, pady=20)
        
        # Navigation buttons with improved styling
        self.copy_button = ctk.CTkButton(self.sidebar, 
                                        text="Manage Excel Copies", 
                                        command=self.show_copy, 
                                        font=ctk.CTkFont(size=16, weight="bold"),
                                        height=50,
                                        fg_color="#2b2b2b",
                                        hover_color="#3a7ebf",
                                        corner_radius=5)
        self.copy_button.grid(row=1, column=0, padx=20, pady=(10, 10), sticky="ew")
        
        self.edit_button = ctk.CTkButton(self.sidebar, 
                                        text="Edit Data", 
                                        command=self.show_edit, 
                                        font=ctk.CTkFont(size=16, weight="bold"),
                                        height=50,
                                        fg_color="#2b2b2b",
                                        hover_color="#3a7ebf",
                                        corner_radius=5)
        self.edit_button.grid(row=2, column=0, padx=20, pady=10, sticky="ew")
        
        # Main content with improved styling
        self.content_frame = ctk.CTkFrame(self, fg_color="#212121")
        self.content_frame.grid(row=0, column=1, sticky="nsew", padx=0, pady=0)
        self.content_frame.grid_columnconfigure(0, weight=1)
        self.content_frame.grid_rowconfigure(0, weight=1)
        
        self.pages = {
            "copy": CopyPage(self.content_frame),
            "edit": EditPage(self.content_frame)
        }
        self.current_page = None
        self.show_copy()
        
        # Appearance mode switcher with improved styling
        appearance_frame = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        appearance_frame.grid(row=9, column=0, sticky="ew", padx=20, pady=(0, 20))
        appearance_frame.grid_columnconfigure(1, weight=1)
        
        ctk.CTkLabel(appearance_frame, 
                    text="Theme:", 
                    font=ctk.CTkFont(size=14)).grid(row=0, column=0, sticky="w", pady=10)
        
        self.appearance_mode_menu = ctk.CTkOptionMenu(
            appearance_frame,
            values=["Light", "Dark", "System"],
            command=self.change_appearance_mode_event,
            width=120,
            font=ctk.CTkFont(size=14)
        )
        self.appearance_mode_menu.set("Dark")
        self.appearance_mode_menu.grid(row=0, column=1, sticky="e", pady=10)
        
    def show_copy(self):
        self._show_page("copy")
        self.copy_button.configure(fg_color="#1f538d")
        self.edit_button.configure(fg_color="#2b2b2b")
        self.pages["copy"].refresh_file_list()
        
    def show_edit(self):
        self._show_page("edit")
        self.edit_button.configure(fg_color="#1f538d")
        self.copy_button.configure(fg_color="#2b2b2b")
        self.pages["edit"].load_files()
        
    def _show_page(self, page):
        if self.current_page:
            self.current_page.grid_remove()
        self.pages[page].grid(row=0, column=0, sticky="nsew")
        self.current_page = self.pages[page]
        
    def change_appearance_mode_event(self, new_appearance_mode: str):
        ctk.set_appearance_mode(new_appearance_mode)

if __name__ == "__main__":
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")
    app = App()
    app.mainloop()
